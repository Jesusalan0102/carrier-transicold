"""
Bitácora de reportes por unidad, agrupada por lote.

Los líderes documentan, por cada unidad de un lote, el trabajo realizado o
los problemas detectados. Al terminar, envían el reporte del lote al
administrador (notificación WebSocket + push) y pueden exportarlo a Excel
para compartirlo, por ejemplo, en el grupo de WhatsApp del trabajo.
"""
import io
import threading
from datetime import date, datetime
from zoneinfo import ZoneInfo
from typing import Optional, Literal

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from db import execute_read, execute_write, execute_write_with_id
from auth import verify_token
from routers.reporte_router import _hdr, _apply, _autofit, AZUL_CORP, AMARILLO, ROJO_CLARO

TZ = ZoneInfo("America/Tijuana")
router = APIRouter(prefix="/api/reportes-unidad", tags=["Reportes de Unidad por Lote"])


def _notify(event: str, payload: dict = None):
    """Emite evento WebSocket+Push desde endpoint síncrono (mismo patrón que
    asignaciones_router.py / tickets_router.py)."""
    def _run():
        import asyncio
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            from routers.ws import notify
            loop.run_until_complete(notify(event, payload))
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"_notify error: {e}")
        finally:
            loop.close()
    threading.Thread(target=_run, daemon=True).start()


def _requiere_lider_o_admin(current_user):
    if current_user["role"] not in ("admin", "lider"):
        raise HTTPException(status_code=403, detail="Solo administradores o líderes")


# ── Modelos ─────────────────────────────────────────────────────────────────
class ReporteUnidadIn(BaseModel):
    id_lote: str
    unit_number: str
    tipo: Literal["trabajo", "problema"]
    detalle: str


class EnviarLoteIn(BaseModel):
    id_lote: str


# ── Selector de lote ─────────────────────────────────────────────────────────
@router.get("/lotes")
def listar_lotes_disponibles(current_user=Depends(verify_token)):
    _requiere_lider_o_admin(current_user)
    return execute_read("""
        SELECT id_lote, COUNT(*) AS total_unidades
        FROM unidades
        WHERE oculto = 0
        GROUP BY id_lote
        ORDER BY id_lote ASC
    """)


# ── Unidades de un lote + borradores del día del líder actual ──────────────
@router.get("/lote/{id_lote}/unidades")
def unidades_del_lote(id_lote: str, current_user=Depends(verify_token)):
    _requiere_lider_o_admin(current_user)
    unidades = execute_read(
        "SELECT unit_number, reefer_model, vin_number FROM unidades "
        "WHERE id_lote=%s AND oculto=0 ORDER BY unit_number",
        (id_lote,)
    )
    if not unidades:
        raise HTTPException(status_code=404, detail=f"No hay unidades registradas en el lote '{id_lote}'")

    hoy = date.today()
    entradas = execute_read(
        "SELECT id, unit_number, tipo, detalle, enviado, username_lider, created_at "
        "FROM reportes_unidad WHERE id_lote=%s AND fecha=%s ORDER BY created_at ASC",
        (id_lote, hoy)
    ) or []

    por_unidad = {}
    for e in entradas:
        por_unidad.setdefault(e["unit_number"], []).append(e)

    for u in unidades:
        u["entradas"] = por_unidad.get(u["unit_number"], [])

    return {"id_lote": id_lote, "fecha": hoy.isoformat(), "unidades": unidades}


# ── Agregar entrada (borrador, aún no enviado) ──────────────────────────────
@router.post("/")
def agregar_entrada(data: ReporteUnidadIn, current_user=Depends(verify_token)):
    _requiere_lider_o_admin(current_user)
    detalle = (data.detalle or "").strip()
    if not detalle:
        raise HTTPException(status_code=400, detail="El detalle no puede estar vacío")

    existe = execute_read(
        "SELECT id FROM unidades WHERE unit_number=%s AND id_lote=%s",
        (data.unit_number, data.id_lote)
    )
    if not existe:
        raise HTTPException(status_code=404, detail="La unidad no pertenece a ese lote")

    nuevo_id = execute_write_with_id(
        "INSERT INTO reportes_unidad (id_lote, unit_number, username_lider, tipo, detalle, fecha) "
        "VALUES (%s, %s, %s, %s, %s, %s)",
        (data.id_lote, data.unit_number, current_user["username"], data.tipo, detalle, date.today())
    )
    return {"mensaje": "Entrada guardada", "id": nuevo_id}


# ── Borrar una entrada propia aún no enviada ────────────────────────────────
@router.delete("/{entrada_id}")
def borrar_entrada(entrada_id: int, current_user=Depends(verify_token)):
    _requiere_lider_o_admin(current_user)
    fila = execute_read("SELECT * FROM reportes_unidad WHERE id=%s", (entrada_id,))
    if not fila:
        raise HTTPException(status_code=404, detail="Entrada no encontrada")
    fila = fila[0]
    if fila["enviado"]:
        raise HTTPException(status_code=400, detail="Esta entrada ya fue enviada al administrador y no se puede borrar")
    if current_user["role"] != "admin" and fila["username_lider"] != current_user["username"]:
        raise HTTPException(status_code=403, detail="Solo puedes borrar tus propias entradas")
    execute_write("DELETE FROM reportes_unidad WHERE id=%s", (entrada_id,))
    return {"mensaje": "Entrada eliminada"}


# ── Enviar el reporte del lote al administrador ─────────────────────────────
@router.post("/enviar")
def enviar_reporte_lote(data: EnviarLoteIn, current_user=Depends(verify_token)):
    _requiere_lider_o_admin(current_user)
    hoy = date.today()

    borradores = execute_read(
        "SELECT id, tipo, unit_number FROM reportes_unidad "
        "WHERE id_lote=%s AND fecha=%s AND username_lider=%s AND enviado=0",
        (data.id_lote, hoy, current_user["username"])
    )
    if not borradores:
        raise HTTPException(status_code=400, detail="No has capturado ninguna entrada de este lote hoy para enviar")

    total_unidades = len({b["unit_number"] for b in borradores})
    total_problemas = sum(1 for b in borradores if b["tipo"] == "problema")

    envio_id = execute_write_with_id(
        "INSERT INTO reportes_lote_envios (id_lote, username_lider, fecha, total_unidades, total_problemas) "
        "VALUES (%s, %s, %s, %s, %s)",
        (data.id_lote, current_user["username"], hoy, total_unidades, total_problemas)
    )
    ids = [b["id"] for b in borradores]
    execute_write(
        f"UPDATE reportes_unidad SET enviado=1, envio_id=%s WHERE id IN ({','.join(['%s'] * len(ids))})",
        tuple([envio_id] + ids)
    )

    nombre_lider_rows = execute_read(
        "SELECT nombre_completo FROM users WHERE username=%s", (current_user["username"],)
    )
    nombre_lider = (nombre_lider_rows[0]["nombre_completo"] if nombre_lider_rows and nombre_lider_rows[0]["nombre_completo"] else current_user["username"])

    admins = execute_read("SELECT username FROM users WHERE role='admin'") or []
    admin_usernames = [a["username"] for a in admins]
    if admin_usernames:
        _notify("reporte_lote_enviado", {
            "usernames": admin_usernames,
            "id_lote": data.id_lote,
            "lider": nombre_lider,
            "total_unidades": total_unidades,
            "total_problemas": total_problemas,
            "envio_id": envio_id,
        })

    return {
        "mensaje": f"Reporte del lote {data.id_lote} enviado al administrador",
        "envio_id": envio_id,
        "total_unidades": total_unidades,
        "total_problemas": total_problemas,
    }


# ── Listar envíos (admin ve todos; líder ve los suyos) ──────────────────────
@router.get("/envios")
def listar_envios(current_user=Depends(verify_token)):
    _requiere_lider_o_admin(current_user)
    if current_user["role"] == "admin":
        filas = execute_read(
            "SELECT * FROM reportes_lote_envios ORDER BY created_at DESC LIMIT 200"
        ) or []
    else:
        filas = execute_read(
            "SELECT * FROM reportes_lote_envios WHERE username_lider=%s ORDER BY created_at DESC LIMIT 200",
            (current_user["username"],)
        ) or []

    nombres = execute_read("SELECT username, nombre_completo FROM users") or []
    mapa = {n["username"]: (n["nombre_completo"] or n["username"]) for n in nombres}
    for f in filas:
        f["nombre_lider"] = mapa.get(f["username_lider"], f["username_lider"])
    return filas


# ── Detalle de un envío (para verlo o volver a exportarlo) ──────────────────
def _detalle_envio(envio_id: int, current_user):
    envio_rows = execute_read("SELECT * FROM reportes_lote_envios WHERE id=%s", (envio_id,))
    if not envio_rows:
        raise HTTPException(status_code=404, detail="Envío no encontrado")
    envio = envio_rows[0]
    if current_user["role"] != "admin" and envio["username_lider"] != current_user["username"]:
        raise HTTPException(status_code=403, detail="No tienes acceso a este envío")

    entradas = execute_read(
        "SELECT unit_number, tipo, detalle, created_at FROM reportes_unidad "
        "WHERE envio_id=%s ORDER BY unit_number, created_at",
        (envio_id,)
    ) or []

    nombres = execute_read("SELECT username, nombre_completo FROM users") or []
    mapa = {n["username"]: (n["nombre_completo"] or n["username"]) for n in nombres}
    envio["nombre_lider"] = mapa.get(envio["username_lider"], envio["username_lider"])
    return envio, entradas


@router.get("/envio/{envio_id}")
def ver_envio(envio_id: int, current_user=Depends(verify_token)):
    _requiere_lider_o_admin(current_user)
    envio, entradas = _detalle_envio(envio_id, current_user)
    return {"envio": envio, "entradas": entradas}


# ── Exportar un envío a Excel (para compartir, ej. en WhatsApp) ────────────
@router.get("/envio/{envio_id}/excel")
def exportar_envio_excel(envio_id: int, current_user=Depends(verify_token)):
    _requiere_lider_o_admin(current_user)
    envio, entradas = _detalle_envio(envio_id, current_user)

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Lote"

    titulo = ws.cell(1, 1, f"REPORTE DE LOTE {envio['id_lote']} — CARRIER TRANSICOLD")
    titulo.font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    titulo.fill = PatternFill("solid", start_color=AZUL_CORP, end_color=AZUL_CORP)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 26

    sub = ws.cell(2, 1,
        f"Líder: {envio['nombre_lider']}   ·   Fecha: {envio['fecha']}   ·   "
        f"Unidades: {envio['total_unidades']}   ·   Problemas: {envio['total_problemas']}"
    )
    sub.font = Font(name="Arial", size=9, italic=True, color="595959")
    ws.merge_cells("A2:D2")
    ws.append([])

    columnas = ["Unidad", "Tipo", "Detalle", "Capturado"]
    ws.append(columnas)
    hstyle = _hdr(AZUL_CORP)
    for col_i in range(1, len(columnas) + 1):
        _apply(ws.cell(4, col_i), hstyle)
    ws.row_dimensions[4].height = 20

    for row_i, e in enumerate(entradas, 5):
        hora = e["created_at"].strftime("%d/%m/%Y %H:%M") if e["created_at"] else ""
        tipo_label = "⚠️ Problema" if e["tipo"] == "problema" else "✅ Trabajo realizado"
        ws.append([e["unit_number"], tipo_label, e["detalle"], hora])
        if e["tipo"] == "problema":
            fill = PatternFill("solid", start_color=ROJO_CLARO, end_color=ROJO_CLARO)
        else:
            fill = PatternFill("solid", start_color=AMARILLO, end_color=AMARILLO) if row_i % 2 == 0 else None
        if fill:
            for col_i in range(1, len(columnas) + 1):
                ws.cell(row_i, col_i).fill = fill
        ws.cell(row_i, 3).alignment = Alignment(wrap_text=True, vertical="top")

    _autofit(ws, min_w=12, max_w=60)
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre_archivo = f"reporte_lote_{envio['id_lote']}_{envio['fecha']}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre_archivo}"'}
    )
