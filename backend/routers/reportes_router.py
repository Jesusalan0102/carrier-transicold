from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import io
from auth import verify_token  # Asegúrate que este import funcione
from db import execute_read

router = APIRouter()

@router.get("/exportar-maestro")
async def exportar_reporte_maestro(token: str = Depends(verify_token)):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Maestro"

        # Cabeceras (ajusta según tus tablas)
        headers = ["ID", "Fecha", "Cliente", "Equipo", "Estado", "Técnico", ...]  # completa según necesites
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Aquí deberías hacer múltiples consultas a las 22 tablas
        # Ejemplo:
        # data = execute_read("SELECT * FROM alguna_tabla LIMIT 1000")
        # for row in data:
        #     ws.append([row['campo1'], row['campo2'], ...])

        # Auto-ajustar columnas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Generar archivo en memoria
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=Reporte_Maestro_Carrier_Transicold.xlsx"
            }
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generando reporte: {str(e)}")
