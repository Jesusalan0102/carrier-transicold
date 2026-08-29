from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from db import execute_read, execute_write, execute_write_with_id
from auth import verify_token
from pydantic import BaseModel
from datetime import datetime
from zoneinfo import ZoneInfo
import io
import logging

TZ = ZoneInfo("America/Tijuana")

# ── Importación opcional de OneDrive ────────────────────────────────────────
try:
    from onedrive_service import sync_reporte_maestro
    ONEDRIVE_ENABLED = True
except ImportError:
    ONEDRIVE_ENABLED = False

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/dashboard", tags=["dashboard"])

ACTIVIDADES_CARRIER = [
    "Cableado","Programación","Soldadura","Check de fugas",
    "Vacío","Cerrado","Pre-viaje","Horas Corridas",
    "Standby","GPS","Corriendo","Inspección",
    "Accesorios","Toma de Valores","Evidencia","Toma de Series",
]

# ── KPIs PRINCIPALES ───────────────────────────────────────────────────────
@router.get("/kpis")
def get_kpis(current_user: dict = Depends(verify_token)):
    total_u = len(execute_read("SELECT id FROM unidades WHERE oculto=0"))
    unidades_visibles = execute_read("SELECT unit_number FROM unidades WHERE oculto=0")
    visibles_set = set(u["unit_number"] for u in unidades_visibles)
    estados_rows = execute_read("SELECT estado, unidad FROM asignaciones")
    estados = [e for e in estados_rows if e["unidad"] in visibles_set]
    completadas = sum(1 for e in estados if e["estado"] == "completada")
    en_proceso  = sum(1 for e in estados if e["estado"] == "en_proceso")
    pendientes  = sum(1 for e in estados if e["estado"] == "pendiente")
    total_t = len(estados)
    avance = round(completadas / total_t * 100) if total_t else 0
    tickets_sin_atender = len(execute_read("SELECT id FROM tickets WHERE atendido=FALSE"))
    solicitudes_pendientes = len(execute_read("SELECT id FROM asignaciones WHERE estado='solicitado'"))
    return {
        "total_unidades": total_u,
        "completadas": completadas,
        "en_proceso": en_proceso,
        "pendientes": pendientes,
        "avance": avance,
        "tickets_sin_atender": tickets_sin_atender,
        "solicitudes_pendientes": solicitudes_pendientes,
    }

# ── STATS POR TÉCNICO (excluye lotes ocultos) ─────────────────────────────
@router.get("/stats_tecnicos")
def get_stats_tecnicos(current_user: dict = Depends(verify_token)):
    return execute_read("""
        SELECT a.tecnico,
               COALESCE(NULLIF(usr.nombre_completo, ''), a.tecnico) as tecnico_display,
               COUNT(*) as total,
               SUM(a.estado='completada') as completadas,
               SUM(a.estado='en_proceso') as en_curso,
               SUM(a.estado='pendiente') as pendientes,
               ROUND(SUM(a.estado='completada') / COUNT(*) * 100) as rendimiento_pct
        FROM asignaciones a
        INNER JOIN unidades u ON u.unit_number = a.unidad
        LEFT JOIN users usr ON usr.username = a.tecnico
        WHERE u.oculto = 0
        GROUP BY a.tecnico, usr.nombre_completo
        ORDER BY completadas DESC
    """)

# ── DISTRIBUCIÓN GLOBAL DE ESTADOS (excluye lotes ocultos) ────────────────
@router.get("/distribucion_global")
def get_distribucion_global(current_user: dict = Depends(verify_token)):
    rows = execute_read("""
        SELECT a.estado, COUNT(*) as total
        FROM asignaciones a
        INNER JOIN unidades u ON u.unit_number = a.unidad
        WHERE u.oculto = 0
        GROUP BY a.estado
    """)
    cnt = {"completada": 0, "en_proceso": 0, "pendiente": 0}
    for r in rows:
        if r["estado"] in cnt:
            cnt[r["estado"]] = r["total"]
    return cnt

# ── ACTIVIDADES OCULTAS DEL DASHBOARD (columnas de la tabla de estatus) ───
@router.get("/actividades_ocultas")
def get_actividades_ocultas(current_user: dict = Depends(verify_token)):
    rows = execute_read("SELECT actividad FROM actividades_ocultas")
    return [r["actividad"] for r in rows]


@router.post("/actividades_ocultas/{actividad}")
def ocultar_actividad(actividad: str, current_user: dict = Depends(verify_token)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores")
    execute_write(
        "INSERT IGNORE INTO actividades_ocultas (actividad) VALUES (%s)",
        (actividad,)
    )
    return {"mensaje": f"Actividad '{actividad}' ocultada del dashboard"}


@router.delete("/actividades_ocultas/{actividad}")
def mostrar_actividad(actividad: str, current_user: dict = Depends(verify_token)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores")
    execute_write(
        "DELETE FROM actividades_ocultas WHERE actividad=%s",
        (actividad,)
    )
    return {"mensaje": f"Actividad '{actividad}' visible nuevamente en el dashboard"}


# ── ESTATUS POR UNIDAD ─────────────────────────────────────────────────────
@router.get("/estatus_unidades")
def get_estatus_unidades(current_user: dict = Depends(verify_token)):
    completadas = execute_read("SELECT unidad, actividad_id FROM asignaciones WHERE estado='completada'")
    completed_set = set((c["unidad"], c["actividad_id"]) for c in completadas)
    unidades = execute_read("SELECT unit_number, id_lote FROM unidades WHERE oculto=0 ORDER BY id_lote, unit_number")
    resultado = []
    for u in unidades:
        row = {"LOTE": u["id_lote"], "#Económico": u["unit_number"]}
        for act in ACTIVIDADES_CARRIER:
            row[act] = "✔" if (u["unit_number"], act) in completed_set else "–"
        resultado.append(row)
    return resultado

# ── CONTADOR ACUMULADO DE HORAS 'CORRIENDO' POR UNIDAD ────────────────────
@router.get("/corriendo_tracking")
def get_corriendo_tracking(current_user: dict = Depends(verify_token)):
    import corriendo_tracking
    return corriendo_tracking.obtener_todos()

# ── DESCARGAR REPORTE EXCEL (+ auto-sync a OneDrive) ──────────────────────
def _generar_excel_maestro_bytes() -> bytes:
    """
    Construye el Excel maestro completo (KPIs, unidades, actividades,
    tickets) y devuelve los bytes .xlsx. Separado del endpoint para
    poder reutilizarlo tanto en la descarga manual (/reporte-excel)
    como en el envío automático del reporte semanal.
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl no instalado. Agrega 'openpyxl' a requirements.txt")

    wb = openpyxl.Workbook()

    # ── Estilos globales ───────────────────────────────────────────────────
    THIN   = Side(style='thin', color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    HDR_FILL  = PatternFill("solid", start_color="1F4E79")
    HDR_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    BODY_FONT = Font(name="Arial", size=9)
    BODY_ALIGN= Alignment(vertical="center", wrap_text=True)

    def write_sheet(ws, rows, col_widths: dict = None):
        if not rows:
            ws.append(["Sin datos"])
            return
        headers = list(rows[0].keys())
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = HDR_ALIGN
            cell.border = BORDER
        for row_idx, row in enumerate(rows, 2):
            for col_idx, val in enumerate(row.values(), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = BODY_FONT
                cell.alignment = BODY_ALIGN
                cell.border = BORDER
        for col_idx, h in enumerate(headers, 1):
            width = (col_widths or {}).get(h, min(len(str(h)) + 6, 35))
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def colorear_estado(ws, rows, col_name):
        if not rows:
            return
        headers = list(rows[0].keys())
        if col_name not in headers:
            return
        col_idx = headers.index(col_name) + 1
        COLOR_MAP = {
            "completada": ("C6EFCE", "276221"),
            "en_proceso": ("DDEBF7", "1F4E79"),
            "pendiente":  ("FFEB9C", "7D6608"),
            "solicitado": ("FFEB9C", "7D6608"),
        }
        for row_idx in range(2, len(rows) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            colors = COLOR_MAP.get(str(cell.value or "").lower())
            if colors:
                cell.fill = PatternFill("solid", start_color=colors[0])
                cell.font = Font(name="Arial", size=9, color=colors[1], bold=True)

    def colorear_semaforo_tickets(ws, rows):
        if not rows:
            return
        headers = list(rows[0].keys())
        n_cols = len(headers)
        atendido_idx  = headers.index("Atendido")        + 1 if "Atendido"        in headers else None
        reporte_idx   = headers.index("Reporte Enviado") + 1 if "Reporte Enviado" in headers else None
        ROJO     = PatternFill("solid", start_color="FFCCCC")
        AMARILLO = PatternFill("solid", start_color="FFFACD")
        VERDE    = PatternFill("solid", start_color="C6EFCE")
        for row_idx, row in enumerate(rows, 2):
            vals = list(row.values())
            atendido    = bool(vals[atendido_idx - 1])  if atendido_idx else False
            reporte_env = bool(vals[reporte_idx - 1])   if reporte_idx  else False
            fill = VERDE if reporte_env else (AMARILLO if atendido else ROJO)
            for col_idx in range(1, n_cols + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    # ── Hoja 1: Resumen KPIs ───────────────────────────────────────────────
    ws_kpi = wb.active
    ws_kpi.title = "Resumen_KPIs"

    total_u_rows = execute_read("SELECT id FROM unidades")
    estados_rows = execute_read("SELECT estado FROM asignaciones")
    comp  = sum(1 for e in estados_rows if e["estado"] == "completada")
    proc  = sum(1 for e in estados_rows if e["estado"] == "en_proceso")
    pend  = sum(1 for e in estados_rows if e["estado"] == "pendiente")
    total = len(estados_rows)
    avance_pct = round(comp / total * 100, 1) if total else 0
    tkt_rows = execute_read("SELECT atendido, reporte_enviado FROM tickets")
    tkt_total     = len(tkt_rows)
    tkt_atendidos = sum(1 for t in tkt_rows if t["atendido"])
    tkt_reporte   = sum(1 for t in tkt_rows if t["reporte_enviado"])

    TITLE_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=13)
    KPI_LABEL   = Font(bold=True, name="Arial", size=10)
    KPI_VAL     = Font(name="Arial", size=12, bold=True, color="1F4E79")
    KPI_ALIGN_R = Alignment(horizontal="right", vertical="center")
    KPI_ALIGN_C = Alignment(horizontal="center", vertical="center")

    ws_kpi.merge_cells("B2:D2")
    ws_kpi["B2"] = "REPORTE MAESTRO — RESUMEN EJECUTIVO"
    ws_kpi["B2"].fill = HDR_FILL
    ws_kpi["B2"].font = TITLE_FONT
    ws_kpi["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws_kpi.row_dimensions[2].height = 30

    kpis = [
        ("Unidades Registradas",        len(total_u_rows)),
        ("Actividades Completadas",     comp),
        ("Actividades En Proceso",      proc),
        ("Actividades Pendientes",      pend),
        ("% Avance General",            f"{avance_pct}%"),
        ("Tickets Totales",             tkt_total),
        ("Tickets Atendidos",           tkt_atendidos),
        ("Reportes de Cierre Enviados", tkt_reporte),
    ]
    for i, (label, val) in enumerate(kpis, 4):
        ws_kpi[f"B{i}"] = label
        ws_kpi[f"B{i}"].font = KPI_LABEL
        ws_kpi[f"B{i}"].alignment = KPI_ALIGN_R
        ws_kpi[f"B{i}"].border = BORDER
        ws_kpi[f"C{i}"] = val
        ws_kpi[f"C{i}"].font = KPI_VAL
        ws_kpi[f"C{i}"].alignment = KPI_ALIGN_C
        ws_kpi[f"C{i}"].border = BORDER
    ws_kpi.column_dimensions["B"].width = 34
    ws_kpi.column_dimensions["C"].width = 18

    # ── Hoja 2: Series Unidades ────────────────────────────────────────────
    ws1 = wb.create_sheet("Series_Unidades")
    unidades = execute_read("SELECT * FROM unidades ORDER BY id_lote, unit_number")
    write_sheet(ws1, unidades, col_widths={
        "unit_number": 14, "id_lote": 14, "vin_number": 20,
        "engine_serial": 20, "compressor_serial": 22,
        "reefer_serial": 20, "reefer_model": 18,
        "evaporator_serial_mjs11": 24, "evaporator_serial_mjd22": 24,
        "generator_serial": 20, "battery_charger_serial": 22,
    })

    # ── Hoja 3: Actividades ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Actividades")
    asigs = execute_read("""
        SELECT a.id, a.unidad, a.actividad_id, a.tecnico, a.estado,
               COALESCE(c.comentarios, a.comentario) AS comentario,
               a.fecha_asignacion, a.fecha_inicio, a.fecha_fin, a.ticket_id
        FROM asignaciones a
        LEFT JOIN (
            SELECT asignacion_id,
                   GROUP_CONCAT(
                       CONCAT(tecnico, ' (', DATE_FORMAT(fecha, '%%d/%%m/%%Y %%H:%%i'), '): ', comentario)
                       ORDER BY fecha SEPARATOR '  |  '
                   ) AS comentarios
            FROM comentarios_actividades
            GROUP BY asignacion_id
        ) c ON c.asignacion_id = a.id
        ORDER BY a.id DESC
    """)
    write_sheet(ws2, asigs, col_widths={
        "actividad_id": 22, "tecnico": 18, "estado": 14,
        "comentario": 50, "fecha_asignacion": 20,
        "fecha_inicio": 20, "fecha_fin": 20,
    })
    colorear_estado(ws2, asigs, "estado")

    # ── Hoja 4: Tickets ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Tickets")
    tickets = execute_read("""
        SELECT
            ticket_num          AS `Ticket #`,
            unit_number         AS `Unidad`,
            vin_number          AS `VIN`,
            descripcion         AS `Problema Reportado`,
            creado_por          AS `Creado Por`,
            tecnico             AS `Técnico Asignado`,
            atendido            AS `Atendido`,
            reporte_enviado     AS `Reporte Enviado`,
            COALESCE(reporte_texto, '—') AS `Reporte Final del Técnico`,
            fecha_creacion      AS `Fecha Creación`,
            fecha_atencion      AS `Fecha Atención`,
            fecha_reporte       AS `Fecha Reporte`
        FROM tickets
        ORDER BY ticket_num DESC
    """)
    write_sheet(ws3, tickets, col_widths={
        "Ticket #": 10, "Unidad": 14, "VIN": 20,
        "Problema Reportado": 35, "Creado Por": 18,
        "Técnico Asignado": 20, "Atendido": 12,
        "Reporte Enviado": 16,
        "Reporte Final del Técnico": 60,
        "Fecha Creación": 20, "Fecha Atención": 20, "Fecha Reporte": 20,
    })
    colorear_semaforo_tickets(ws3, tickets)

    # ── Hoja 5: Reporte Cierre ─────────────────────────────────────────────
    ws4 = wb.create_sheet("Reporte_Cierre_Tickets")
    cierre = execute_read("""
        SELECT
            t.ticket_num                        AS `Ticket #`,
            t.unit_number                       AS `Unidad`,
            t.vin_number                        AS `VIN`,
            t.descripcion                       AS `Problema Reportado`,
            t.creado_por                        AS `Creado Por`,
            t.tecnico                           AS `Técnico Asignado`,
            t.fecha_creacion                    AS `Fecha Creación`,
            t.fecha_atencion                    AS `Fecha Atención`,
            a.actividad_id                      AS `Actividad`,
            a.estado                            AS `Estado Actividad`,
            a.fecha_inicio                      AS `Inicio Trabajo`,
            a.fecha_fin                         AS `Fin Trabajo`,
            a.comentario                        AS `Comentario Técnico`,
            COALESCE(t.reporte_texto, '—')      AS `Reporte Final del Técnico`,
            t.fecha_reporte                     AS `Fecha Reporte Final`,
            t.reporte_enviado                   AS `Ticket Cerrado`
        FROM tickets t
        LEFT JOIN asignaciones a ON a.ticket_id = t.id
        ORDER BY t.ticket_num DESC
    """)
    write_sheet(ws4, cierre, col_widths={
        "Ticket #": 10, "Unidad": 12, "VIN": 20,
        "Problema Reportado": 35, "Creado Por": 18,
        "Técnico Asignado": 20, "Fecha Creación": 20,
        "Fecha Atención": 20, "Actividad": 22,
        "Estado Actividad": 16, "Inicio Trabajo": 20,
        "Fin Trabajo": 20, "Comentario Técnico": 45,
        "Reporte Final del Técnico": 60,
        "Fecha Reporte Final": 22, "Ticket Cerrado": 15,
    })
    colorear_estado(ws4, cierre, "Estado Actividad")

    if cierre:
        headers_cierre = list(cierre[0].keys())
        if "Reporte Final del Técnico" in headers_cierre:
            col_rf = headers_cierre.index("Reporte Final del Técnico") + 1
            RF_FILL = PatternFill("solid", start_color="EEF4FB")
            RF_FONT = Font(name="Arial", size=9, color="1F4E79", italic=True)
            for row_idx in range(2, len(cierre) + 2):
                cell = ws4.cell(row=row_idx, column=col_rf)
                cell.fill = RF_FILL
                cell.font = RF_FONT
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    # ── Generar bytes del Excel ────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    excel_bytes = buf.getvalue()
    return excel_bytes


@router.get("/reporte-excel")
def reporte_excel(current_user: dict = Depends(verify_token)):
    excel_bytes = _generar_excel_maestro_bytes()

    fecha = datetime.now(TZ).strftime("%Y-%m-%d")

    # ── Auto-sync a OneDrive cada vez que se descarga el reporte ──────────
    if ONEDRIVE_ENABLED:
        try:
            web_url = sync_reporte_maestro(excel_bytes, fecha)
            logger.info(f"[OneDrive] Reporte guardado: {web_url}")
        except Exception as e:
            # No bloquear la descarga si OneDrive falla
            logger.warning(f"[OneDrive] No se pudo sincronizar reporte: {e}")

    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Carrier_Reporte_{fecha}.xlsx"}
    )


# ── KPIs POR TÉCNICO (vista ejecutiva consolidada) ─────────────────────────
# A diferencia de /stats_tecnicos (que solo cuenta actividades por estado),
# esto cruza tickets + asistencia para dar una foto más completa de cada
# técnico: qué tan rápido cierra tickets, si adjunta reporte, y su
# puntualidad. Solo admin/líder puede verlo — es información de desempeño.
@router.get("/kpis_tecnico")
def get_kpis_tecnico(dias: int = 30, current_user: dict = Depends(verify_token)):
    if current_user["role"] not in ("admin", "lider"):
        raise HTTPException(status_code=403, detail="Solo administradores y líderes")

    if dias < 1 or dias > 365:
        raise HTTPException(status_code=400, detail="'dias' debe estar entre 1 y 365")

    # 1. Tiempo promedio de resolución + tickets con/sin reporte adjunto,
    #    por técnico. Se basa en asignaciones.tecnico (quien atendió el
    #    ticket), no en tickets.creado_por (quien lo levantó).
    tickets_por_tecnico = execute_read(
        """
        SELECT
            a.tecnico,
            COALESCE(NULLIF(usr.nombre_completo, ''), a.tecnico) AS tecnico_display,
            COUNT(DISTINCT t.id) AS tickets_cerrados,
            SUM(t.reporte_archivo_nombre IS NOT NULL) AS con_reporte_adjunto,
            SUM(t.reporte_archivo_nombre IS NULL) AS sin_reporte_adjunto,
            ROUND(AVG(TIMESTAMPDIFF(MINUTE, t.fecha_creacion, t.fecha_reporte)) / 60, 1)
                AS horas_promedio_resolucion
        FROM tickets t
        INNER JOIN asignaciones a ON a.ticket_id = t.id
        LEFT JOIN users usr ON usr.username = a.tecnico
        WHERE t.reporte_enviado = TRUE
          AND t.fecha_reporte >= DATE_SUB(NOW(), INTERVAL %s DAY)
        GROUP BY a.tecnico, usr.nombre_completo
        ORDER BY tickets_cerrados DESC
        """,
        (dias,)
    )

    # 2. Asistencia vs. tardanzas por técnico en la misma ventana de días.
    #    retardo_min > 0 marca un check-in tarde (ver migración en db.py).
    asistencia_por_tecnico = execute_read(
        """
        SELECT
            ra.username AS tecnico,
            COALESCE(NULLIF(usr.nombre_completo, ''), ra.username) AS tecnico_display,
            COUNT(*) AS dias_con_checkin,
            SUM(ra.retardo_min > 0) AS dias_con_tardanza,
            ROUND(AVG(ra.retardo_min), 1) AS retardo_promedio_min,
            ROUND(SUM(ra.retardo_min > 0) / COUNT(*) * 100) AS pct_tardanza
        FROM registros_asistencia ra
        LEFT JOIN users usr ON usr.username = ra.username
        WHERE ra.tipo = 'entrada'
          AND ra.fecha >= DATE_SUB(CURDATE(), INTERVAL %s DAY)
        GROUP BY ra.username, usr.nombre_completo
        ORDER BY pct_tardanza DESC
        """,
        (dias,)
    )

    # 3. Merge por técnico: cada técnico puede tener datos de tickets,
    #    de asistencia, o ambos — no asumimos que siempre existan los dos.
    por_tecnico = {}
    for row in tickets_por_tecnico:
        por_tecnico[row["tecnico"]] = {
            "tecnico": row["tecnico"],
            "tecnico_display": row["tecnico_display"],
            "tickets_cerrados": row["tickets_cerrados"],
            "con_reporte_adjunto": row["con_reporte_adjunto"],
            "sin_reporte_adjunto": row["sin_reporte_adjunto"],
            "horas_promedio_resolucion": row["horas_promedio_resolucion"],
            "dias_con_checkin": None,
            "dias_con_tardanza": None,
            "retardo_promedio_min": None,
            "pct_tardanza": None,
        }
    for row in asistencia_por_tecnico:
        entry = por_tecnico.setdefault(row["tecnico"], {
            "tecnico": row["tecnico"],
            "tecnico_display": row["tecnico_display"],
            "tickets_cerrados": 0,
            "con_reporte_adjunto": 0,
            "sin_reporte_adjunto": 0,
            "horas_promedio_resolucion": None,
        })
        entry["dias_con_checkin"] = row["dias_con_checkin"]
        entry["dias_con_tardanza"] = row["dias_con_tardanza"]
        entry["retardo_promedio_min"] = row["retardo_promedio_min"]
        entry["pct_tardanza"] = row["pct_tardanza"]

    return {
        "dias": dias,
        "tecnicos": sorted(
            por_tecnico.values(),
            key=lambda x: x["tickets_cerrados"],
            reverse=True,
        ),
    }


# ── REPORTE SEMANAL AUTOMÁTICO (envío manual / prueba) ─────────────────────
# El envío automático corre solo (ver reportes_semanales.programador_
# reporte_semanal, disparado desde main.py). Este endpoint es para que un
# admin pueda probar la configuración (destinatarios, permiso Mail.Send de
# Graph, etc.) sin tener que esperar al día/hora programados.
@router.post("/reporte-semanal/enviar-ahora")
def enviar_reporte_semanal_ahora(current_user: dict = Depends(verify_token)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores")

    from reportes_semanales import enviar_reporte_semanal
    resultado = enviar_reporte_semanal(forzar=True)
    if not resultado["enviado"]:
        raise HTTPException(status_code=400, detail=resultado["motivo"])
    return resultado


# ── KPIs PERSONALIZADOS (definidos por el admin) ────────────────────────────
# Complementa /kpis_tecnico: además de lo que el sistema ya calcula solo
# (tickets, asistencia), un admin puede definir sus propias métricas
# ("Satisfacción de cliente", "Unidades PDI completadas", lo que sea) y
# capturar el valor a mano por técnico y por periodo.
class KpiMetricaCreate(BaseModel):
    nombre: str
    unidad: str = ""
    descripcion: str = ""
    tipo_evaluacion: str = "manual_directo"  # manual_directo | manual_rango | manual_rango_invertido
    valor_min: float = 0
    valor_max: float = 100
    peso: float = 0


class KpiMetricaConfig(BaseModel):
    """Para editar una métrica ya creada (automática o manual)."""
    tipo_evaluacion: str | None = None
    valor_min: float | None = None
    valor_max: float | None = None
    peso: float


class KpiValorUpsert(BaseModel):
    metrica_id: int
    tecnico: str
    periodo: str  # texto libre, ej. "2026-08" (mes) o "2026-W35" (semana ISO)
    valor: float | None = None


TIPOS_EVALUACION_VALIDOS = {"manual_directo", "manual_rango", "manual_rango_invertido"}


@router.get("/kpis_custom/metricas")
def listar_kpi_metricas(current_user: dict = Depends(verify_token)):
    if current_user["role"] not in ("admin", "lider"):
        raise HTTPException(status_code=403, detail="Solo administradores y líderes")
    return execute_read(
        "SELECT id, nombre, unidad, descripcion, activo, tipo_evaluacion, "
        "valor_min, valor_max, peso, es_automatica, clave_automatica "
        "FROM kpis_custom_metricas WHERE activo = 1 ORDER BY es_automatica DESC, nombre"
    )


@router.post("/kpis_custom/metricas")
def crear_kpi_metrica(data: KpiMetricaCreate, current_user: dict = Depends(verify_token)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores")
    nombre = data.nombre.strip()
    if not nombre:
        raise HTTPException(status_code=400, detail="El nombre de la métrica no puede estar vacío")
    if data.tipo_evaluacion not in TIPOS_EVALUACION_VALIDOS:
        raise HTTPException(status_code=400, detail=f"tipo_evaluacion debe ser uno de {TIPOS_EVALUACION_VALIDOS}")
    if data.tipo_evaluacion != "manual_directo" and data.valor_max <= data.valor_min:
        raise HTTPException(status_code=400, detail="valor_max debe ser mayor que valor_min")
    if not (0 <= data.peso <= 100):
        raise HTTPException(status_code=400, detail="El peso debe estar entre 0 y 100")

    nuevo_id = execute_write_with_id(
        "INSERT INTO kpis_custom_metricas "
        "(nombre, unidad, descripcion, creado_por, tipo_evaluacion, valor_min, valor_max, peso) "
        "VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
        (nombre, data.unidad.strip(), data.descripcion.strip(), current_user["username"],
         data.tipo_evaluacion, data.valor_min, data.valor_max, data.peso)
    )
    return {"id": nuevo_id, "nombre": nombre}


@router.put("/kpis_custom/metricas/{metrica_id}/config")
def configurar_kpi_metrica(metrica_id: int, data: KpiMetricaConfig, current_user: dict = Depends(verify_token)):
    """
    Ajusta el peso (y, para métricas manuales, el tipo de evaluación y sus
    márgenes) de una métrica ya existente. Para las 3 métricas automáticas
    (tiempo SLA, reporte, asistencia) su lógica de cálculo está fija en
    código — aquí solo se les puede cambiar el peso.
    """
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores")
    if not (0 <= data.peso <= 100):
        raise HTTPException(status_code=400, detail="El peso debe estar entre 0 y 100")

    filas = execute_read(
        "SELECT es_automatica FROM kpis_custom_metricas WHERE id = %s AND activo = 1", (metrica_id,)
    )
    if not filas:
        raise HTTPException(status_code=404, detail="Métrica no encontrada")

    if filas[0]["es_automatica"]:
        execute_write("UPDATE kpis_custom_metricas SET peso = %s WHERE id = %s", (data.peso, metrica_id))
        return {"ok": True}

    tipo_evaluacion = data.tipo_evaluacion or "manual_directo"
    valor_min = data.valor_min if data.valor_min is not None else 0
    valor_max = data.valor_max if data.valor_max is not None else 100
    if tipo_evaluacion not in TIPOS_EVALUACION_VALIDOS:
        raise HTTPException(status_code=400, detail=f"tipo_evaluacion debe ser uno de {TIPOS_EVALUACION_VALIDOS}")
    if tipo_evaluacion != "manual_directo" and valor_max <= valor_min:
        raise HTTPException(status_code=400, detail="valor_max debe ser mayor que valor_min")

    execute_write(
        "UPDATE kpis_custom_metricas SET tipo_evaluacion=%s, valor_min=%s, valor_max=%s, peso=%s WHERE id=%s",
        (tipo_evaluacion, valor_min, valor_max, data.peso, metrica_id)
    )
    return {"ok": True}


@router.delete("/kpis_custom/metricas/{metrica_id}")
def desactivar_kpi_metrica(metrica_id: int, current_user: dict = Depends(verify_token)):
    """
    Desactiva la métrica (no la borra) para no perder el histórico de
    valores ya capturados — simplemente deja de aparecer en la tabla.
    """
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores")
    filas = execute_read(
        "SELECT id, es_automatica FROM kpis_custom_metricas WHERE id = %s", (metrica_id,)
    )
    if not filas:
        raise HTTPException(status_code=404, detail="Métrica no encontrada")
    if filas[0]["es_automatica"]:
        raise HTTPException(
            status_code=400,
            detail="Las métricas automáticas no se pueden quitar, solo ponerles peso 0 para que no cuenten"
        )
    execute_write("UPDATE kpis_custom_metricas SET activo = 0 WHERE id = %s", (metrica_id,))
    return {"ok": True}


@router.get("/kpis_custom/valores")
def listar_kpi_valores(periodo: str = Query(...), current_user: dict = Depends(verify_token)):
    if current_user["role"] not in ("admin", "lider"):
        raise HTTPException(status_code=403, detail="Solo administradores y líderes")

    metricas = execute_read(
        "SELECT id, nombre, unidad, descripcion, tipo_evaluacion, valor_min, valor_max, peso, "
        "es_automatica, clave_automatica FROM kpis_custom_metricas "
        "WHERE activo = 1 ORDER BY es_automatica DESC, nombre"
    )
    tecnicos = execute_read(
        "SELECT username, COALESCE(NULLIF(nombre_completo, ''), username) AS nombre_display "
        "FROM users WHERE role = 'tecnico' ORDER BY nombre_display"
    )
    valores = execute_read(
        "SELECT metrica_id, tecnico, valor FROM kpis_custom_valores WHERE periodo = %s",
        (periodo,)
    )
    mapa_valores = {(v["metrica_id"], v["tecnico"]): v["valor"] for v in valores}

    # Las métricas automáticas no se editan a mano — se muestran en esta
    # vista solo como referencia de que existen y cuánto peso tienen.
    metricas_manuales = [m for m in metricas if not m["es_automatica"]]

    return {
        "periodo": periodo,
        "metricas": metricas,
        "tecnicos": tecnicos,
        "valores": [
            {"metrica_id": m["id"], "tecnico": t["username"], "valor": mapa_valores.get((m["id"], t["username"]))}
            for m in metricas_manuales
            for t in tecnicos
        ],
    }


@router.put("/kpis_custom/valores")
def guardar_kpi_valor(data: KpiValorUpsert, current_user: dict = Depends(verify_token)):
    if current_user["role"] not in ("admin", "lider"):
        raise HTTPException(status_code=403, detail="Solo administradores y líderes")

    periodo = data.periodo.strip()
    if not periodo:
        raise HTTPException(status_code=400, detail="El periodo no puede estar vacío")

    metrica = execute_read(
        "SELECT id FROM kpis_custom_metricas WHERE id = %s AND activo = 1", (data.metrica_id,)
    )
    if not metrica:
        raise HTTPException(status_code=404, detail="Métrica no encontrada o desactivada")

    execute_write(
        """
        INSERT INTO kpis_custom_valores (metrica_id, tecnico, periodo, valor, registrado_por)
        VALUES (%s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE valor = %s, registrado_por = %s
        """,
        (data.metrica_id, data.tecnico, periodo, data.valor,
         current_user["username"], data.valor, current_user["username"])
    )
    return {"ok": True}


# ── SCORE FINAL PONDERADO (0-100) ───────────────────────────────────────────
def _score_manual(valor, tipo_evaluacion, valor_min, valor_max):
    if valor is None:
        return None
    if tipo_evaluacion == "manual_directo":
        return max(0.0, min(100.0, float(valor)))
    rango = float(valor_max) - float(valor_min)
    if rango <= 0:
        return None
    pct = (float(valor) - float(valor_min)) / rango * 100
    if tipo_evaluacion == "manual_rango_invertido":
        pct = 100 - pct
    return max(0.0, min(100.0, pct))


@router.get("/kpis_score")
def kpi_score_final(dias: int = 30, periodo: str = Query(...), current_user: dict = Depends(verify_token)):
    """
    KPI final 0-100 por técnico: junta las 3 métricas automáticas (tiempo
    SLA, reporte adjunto, asistencia — calculadas en vivo sobre `dias`) con
    todas las métricas manuales que tengan un valor capturado en `periodo`,
    y las pondera según el peso que el admin le puso a cada una.
    """
    if current_user["role"] not in ("admin", "lider"):
        raise HTTPException(status_code=403, detail="Solo administradores y líderes")

    metricas = execute_read(
        "SELECT id, nombre, unidad, tipo_evaluacion, valor_min, valor_max, peso, "
        "es_automatica, clave_automatica FROM kpis_custom_metricas WHERE activo = 1"
    )
    metricas_con_peso = [m for m in metricas if m["peso"] and m["peso"] > 0]
    suma_pesos = round(sum(m["peso"] for m in metricas_con_peso), 2)

    tecnicos = execute_read(
        "SELECT username, COALESCE(NULLIF(nombre_completo, ''), username) AS nombre_display "
        "FROM users WHERE role = 'tecnico' ORDER BY nombre_display"
    )
    if not tecnicos:
        return {"periodo": periodo, "dias": dias, "suma_pesos": suma_pesos, "tecnicos": []}

    # ── Valores crudos de las 3 métricas automáticas, por técnico ──────────
    filas_sla = execute_read(
        """
        SELECT tecnico,
               TIMESTAMPDIFF(MINUTE, fecha_asignacion, fecha_fin) / 60.0 AS horas_reales,
               tiempo_estimado_horas
        FROM asignaciones
        WHERE estado = 'completada'
          AND tiempo_estimado_horas IS NOT NULL
          AND fecha_asignacion IS NOT NULL
          AND fecha_fin >= DATE_SUB(NOW(), INTERVAL %s DAY)
        """,
        (dias,)
    )
    scores_sla_por_tecnico = {}
    for f in filas_sla:
        objetivo = float(f["tiempo_estimado_horas"])
        real = float(f["horas_reales"])
        if objetivo <= 0:
            continue
        score = 100.0 if real <= objetivo else max(0.0, 100 - (real - objetivo) / objetivo * 100)
        scores_sla_por_tecnico.setdefault(f["tecnico"], []).append(score)
    scores_sla_por_tecnico = {t: sum(v) / len(v) for t, v in scores_sla_por_tecnico.items()}

    filas_reporte = execute_read(
        """
        SELECT a.tecnico,
               SUM(t.reporte_archivo_nombre IS NOT NULL) * 100.0 / COUNT(*) AS pct
        FROM tickets t
        INNER JOIN asignaciones a ON a.ticket_id = t.id
        WHERE t.reporte_enviado = TRUE
          AND t.fecha_reporte >= DATE_SUB(NOW(), INTERVAL %s DAY)
        GROUP BY a.tecnico
        """,
        (dias,)
    )
    scores_reporte_por_tecnico = {f["tecnico"]: float(f["pct"]) for f in filas_reporte}

    filas_asistencia = execute_read(
        """
        SELECT ra.username AS tecnico,
               100 - (SUM(ra.retardo_min > 0) * 100.0 / COUNT(*)) AS score
        FROM registros_asistencia ra
        WHERE ra.tipo = 'entrada'
          AND ra.fecha >= DATE_SUB(CURDATE(), INTERVAL %s DAY)
        GROUP BY ra.username
        """,
        (dias,)
    )
    scores_asistencia_por_tecnico = {f["tecnico"]: float(f["score"]) for f in filas_asistencia}

    # ── Calidad / Retrabajos: penaliza a quien hizo el trabajo original ────
    # de una unidad si después se completa un "Retrabajo Eléctrico" o
    # "Retrabajo Soldador" sobre esa misma unidad. La ventana `dias` limita
    # qué retrabajos cuentan y cuánto trabajo original reciente se usa como
    # base de comparación; a quién se le atribuye el retrabajo se busca en
    # todo el historial (el cableado pudo haberse hecho antes de la ventana).
    ELECTRICO_ACTS = ("Cableado", "Programación", "Extra Eléctrico")
    SOLDADOR_ACTS = ("Soldadura", "Check de fugas", "Extra Soldador")

    def _unidades_trabajadas_por_tecnico(actividades):
        ph = ",".join(["%s"] * len(actividades))
        filas = execute_read(
            f"SELECT tecnico, COUNT(DISTINCT unidad) AS c FROM asignaciones "
            f"WHERE estado='completada' AND actividad_id IN ({ph}) "
            f"AND fecha_fin >= DATE_SUB(NOW(), INTERVAL %s DAY) GROUP BY tecnico",
            tuple(list(actividades) + [dias])
        ) or []
        return {f["tecnico"]: f["c"] for f in filas}

    unidades_electrico_por_tecnico = _unidades_trabajadas_por_tecnico(ELECTRICO_ACTS)
    unidades_soldador_por_tecnico = _unidades_trabajadas_por_tecnico(SOLDADOR_ACTS)

    def _responsables_originales(unidad, actividades):
        ph = ",".join(["%s"] * len(actividades))
        filas = execute_read(
            f"SELECT DISTINCT tecnico FROM asignaciones WHERE unidad=%s AND estado='completada' "
            f"AND actividad_id IN ({ph})",
            tuple([unidad] + list(actividades))
        ) or []
        return [f["tecnico"] for f in filas]

    def _retrabajos_atribuidos(actividad_retrabajo, actividades_originales):
        filas = execute_read(
            "SELECT unidad FROM asignaciones WHERE estado='completada' AND actividad_id=%s "
            "AND fecha_fin >= DATE_SUB(NOW(), INTERVAL %s DAY)",
            (actividad_retrabajo, dias)
        ) or []
        atribuidos = {}
        for f in filas:
            for t in _responsables_originales(f["unidad"], actividades_originales):
                atribuidos[t] = atribuidos.get(t, 0) + 1
        return atribuidos

    retrabajos_electrico_por_tecnico = _retrabajos_atribuidos("Retrabajo Eléctrico", ELECTRICO_ACTS)
    retrabajos_soldador_por_tecnico = _retrabajos_atribuidos("Retrabajo Soldador", SOLDADOR_ACTS)

    scores_calidad_por_tecnico = {}
    valores_crudos_calidad = {}
    for t in tecnicos:
        u = t["username"]
        unidades = unidades_electrico_por_tecnico.get(u, 0) + unidades_soldador_por_tecnico.get(u, 0)
        retrabajos = retrabajos_electrico_por_tecnico.get(u, 0) + retrabajos_soldador_por_tecnico.get(u, 0)
        if unidades == 0 and retrabajos == 0:
            continue  # sin trabajo original ni retrabajos en la ventana — sin dato
        base = max(unidades, 1)
        scores_calidad_por_tecnico[u] = max(0.0, 100 - (retrabajos / base * 100))
        valores_crudos_calidad[u] = retrabajos

    # ── Valores manuales del periodo ────────────────────────────────────────
    valores_manuales = execute_read(
        "SELECT metrica_id, tecnico, valor FROM kpis_custom_valores WHERE periodo = %s", (periodo,)
    )
    mapa_manuales = {(v["metrica_id"], v["tecnico"]): v["valor"] for v in valores_manuales}

    resultado_tecnicos = []
    for t in tecnicos:
        username = t["username"]
        detalle = []
        suma_ponderada = 0.0
        peso_aplicado = 0.0

        for m in metricas_con_peso:
            if m["clave_automatica"] == "automatica_tiempo_sla":
                valor_crudo = scores_sla_por_tecnico.get(username)
                score = valor_crudo  # el score YA es 0-100 (se calculó por asignación)
            elif m["clave_automatica"] == "automatica_reporte":
                valor_crudo = scores_reporte_por_tecnico.get(username)
                score = valor_crudo
            elif m["clave_automatica"] == "automatica_asistencia":
                valor_crudo = scores_asistencia_por_tecnico.get(username)
                score = valor_crudo
            elif m["clave_automatica"] == "automatica_calidad_retrabajo":
                valor_crudo = valores_crudos_calidad.get(username)
                score = scores_calidad_por_tecnico.get(username)
            else:
                valor_crudo = mapa_manuales.get((m["id"], username))
                score = _score_manual(valor_crudo, m["tipo_evaluacion"], m["valor_min"], m["valor_max"])

            if score is None:
                continue  # sin dato para este técnico en esta métrica — no cuenta, ni resta

            aporte = score * float(m["peso"]) / 100
            suma_ponderada += aporte
            peso_aplicado += float(m["peso"])
            detalle.append({
                "metrica": m["nombre"], "unidad": m["unidad"], "peso": float(m["peso"]),
                "valor_crudo": round(float(valor_crudo), 2), "score": round(float(score), 1),
                "aporte": round(aporte, 2),
            })

        # Se re-normaliza sobre el peso que SÍ tuvo dato, para que a un
        # técnico nuevo (sin historial en alguna métrica) no se le castigue
        # con puntos en cero por falta de datos en vez de falta de desempeño.
        score_final = round(suma_ponderada / peso_aplicado * 100, 1) if peso_aplicado > 0 else None

        resultado_tecnicos.append({
            "tecnico": username,
            "tecnico_display": t["nombre_display"],
            "score_final": score_final,
            "peso_con_datos": round(peso_aplicado, 2),
            "detalle": detalle,
        })

    resultado_tecnicos.sort(key=lambda x: (x["score_final"] is None, -(x["score_final"] or 0)))

    return {
        "periodo": periodo,
        "dias": dias,
        "suma_pesos": suma_pesos,
        "advertencia": None if suma_pesos == 100 else
            f"Los pesos de las métricas activas suman {suma_pesos}%, no 100% — el score se re-normaliza pero revisa tu configuración.",
        "tecnicos": resultado_tecnicos,
    }


# ── Exportar el score de KPIs a Excel, con el desglose de cómo se calculó ──
@router.get("/kpis_score/exportar")
def exportar_kpi_score(dias: int = 30, periodo: str = Query(...), current_user: dict = Depends(verify_token)):
    if current_user["role"] not in ("admin", "lider"):
        raise HTTPException(status_code=403, detail="Solo administradores y líderes")

    data = kpi_score_final(dias=dias, periodo=periodo, current_user=current_user)

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        raise RuntimeError("openpyxl no instalado. Agrega 'openpyxl' a requirements.txt")

    THIN = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    HDR_FILL = PatternFill("solid", start_color="1F4E79")
    HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    BODY_FONT = Font(name="Arial", size=9)
    TITLE_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=14)

    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen — score final por técnico ───────────────────────────
    ws = wb.active
    ws.title = "Resumen"
    ws.merge_cells("A1:D1")
    ws["A1"] = f"KPI SCORE — Periodo {data['periodo']}  ·  ventana automática: últimos {data['dias']} días"
    ws["A1"].fill = HDR_FILL
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = HDR_ALIGN
    ws.row_dimensions[1].height = 28

    if data.get("advertencia"):
        ws.merge_cells("A2:D2")
        ws["A2"] = f"⚠️ {data['advertencia']}"
        ws["A2"].font = Font(italic=True, color="B45309", name="Arial", size=9)

    fila = 4
    encabezados = ["Técnico", "Score final (0-100)", "% de peso con datos", "Métricas evaluadas"]
    for col, h in enumerate(encabezados, 1):
        c = ws.cell(fila, col, h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = HDR_ALIGN
        c.border = BORDER
    fila += 1
    for t in data["tecnicos"]:
        ws.cell(fila, 1, t["tecnico_display"]).font = BODY_FONT
        cscore = ws.cell(fila, 2, t["score_final"] if t["score_final"] is not None else "Sin datos")
        cscore.font = Font(bold=True, name="Arial", size=11, color=(
            "1A7A4A" if (t["score_final"] or 0) >= 90 else
            "B45309" if (t["score_final"] or 0) >= 70 else "C0392B"
        ) if t["score_final"] is not None else "808080")
        cscore.alignment = Alignment(horizontal="center")
        ws.cell(fila, 3, f"{t['peso_con_datos']}%").alignment = Alignment(horizontal="center")
        ws.cell(fila, 4, len(t["detalle"])).alignment = Alignment(horizontal="center")
        for col in range(1, 5):
            ws.cell(fila, col).border = BORDER
        fila += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    # ── Hoja 2: Desglose — cómo se calculó cada score, métrica por métrica ──
    ws2 = wb.create_sheet("Desglose del cálculo")
    ws2.merge_cells("A1:F1")
    ws2["A1"] = "CÓMO SE OBTUVO CADA SCORE (métrica por métrica)"
    ws2["A1"].fill = HDR_FILL
    ws2["A1"].font = TITLE_FONT
    ws2["A1"].alignment = HDR_ALIGN
    ws2.row_dimensions[1].height = 26

    fila2 = 3
    encabezados2 = ["Técnico", "Métrica", "Valor crudo", "Unidad", "Score (0-100)", "Peso", "Aporte al score final"]
    for col, h in enumerate(encabezados2, 1):
        c = ws2.cell(fila2, col, h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = HDR_ALIGN
        c.border = BORDER
    fila2 += 1
    for t in data["tecnicos"]:
        if not t["detalle"]:
            ws2.cell(fila2, 1, t["tecnico_display"]).font = BODY_FONT
            ws2.cell(fila2, 2, "Sin métricas con dato en este periodo/ventana").font = Font(italic=True, color="808080", name="Arial", size=9)
            for col in range(1, 8):
                ws2.cell(fila2, col).border = BORDER
            fila2 += 1
            continue
        for d in t["detalle"]:
            ws2.cell(fila2, 1, t["tecnico_display"]).font = BODY_FONT
            ws2.cell(fila2, 2, d["metrica"]).font = BODY_FONT
            ws2.cell(fila2, 3, d["valor_crudo"]).font = BODY_FONT
            ws2.cell(fila2, 4, d["unidad"] or "").font = BODY_FONT
            ws2.cell(fila2, 5, d["score"]).font = BODY_FONT
            ws2.cell(fila2, 6, f"{d['peso']}%").font = BODY_FONT
            ws2.cell(fila2, 7, d["aporte"]).font = BODY_FONT
            for col in range(1, 8):
                ws2.cell(fila2, col).border = BORDER
                ws2.cell(fila2, col).alignment = Alignment(vertical="center")
            fila2 += 1
    for col, w in zip("ABCDEFG", [26, 26, 14, 12, 14, 10, 18]):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A4"

    # ── Hoja 3: Metodología — explicación de cómo se calcula cada métrica ──
    ws3 = wb.create_sheet("Metodología")
    ws3.merge_cells("A1:B1")
    ws3["A1"] = "METODOLOGÍA DE CÁLCULO"
    ws3["A1"].fill = HDR_FILL
    ws3["A1"].font = TITLE_FONT
    ws3["A1"].alignment = HDR_ALIGN
    ws3.row_dimensions[1].height = 26

    explicaciones = [
        ("Tiempo / SLA (automática)",
         f"Por cada actividad completada en los últimos {dias} días, se compara el tiempo real contra el tiempo "
         "estimado. Si el técnico terminó a tiempo o antes, esa actividad vale 100 puntos; si se pasó, pierde "
         "puntos proporcionalmente al porcentaje que se excedió. El score de la métrica es el promedio de todas "
         "sus actividades en la ventana."),
        ("Reporte adjunto (automática)",
         f"De los tickets con reporte enviado en los últimos {dias} días, qué porcentaje además tiene un archivo "
         "de reporte adjunto. 100% = siempre adjuntó archivo."),
        ("Asistencia (automática)",
         f"De los checados de entrada en los últimos {dias} días, qué porcentaje NO tuvo retardo. "
         "100% = nunca llegó tarde."),
        ("Calidad / Retrabajos (automática)",
         f"Si en los últimos {dias} días se completa un 'Retrabajo Eléctrico' o 'Retrabajo Soldador' sobre una "
         "unidad, se busca en todo el historial quién hizo el trabajo original de esa especialidad en esa "
         "unidad (Cableado/Programación/Extra Eléctrico para eléctrico; Soldadura/Check de fugas/Extra "
         "Soldador para soldador) y se le atribuye el retrabajo. El score es 100 menos el % de retrabajos "
         "atribuidos sobre las unidades que ese técnico trabajó en la ventana (mínimo 1 unidad para no dividir "
         "entre cero)."),
        ("Métricas manuales",
         "Capturadas a mano por el admin/líder para el periodo seleccionado. 'Directo' usa el valor tal cual "
         "(0-100). 'Rango' convierte el valor a un porcentaje entre un mínimo y un máximo configurados. "
         "'Rango invertido' hace lo mismo pero invertido (útil cuando un valor más bajo es mejor, ej. quejas)."),
        ("Score final ponderado",
         "Cada métrica aporta (score de la métrica × su peso%) / 100. Se suman los aportes de las métricas que "
         "SÍ tuvieron dato para ese técnico, y se dividen entre la suma de esos pesos (no entre 100), para no "
         "castigar a un técnico por falta de datos en una métrica en vez de por desempeño."),
    ]
    fila3 = 3
    for titulo, texto in explicaciones:
        ws3.cell(fila3, 1, titulo).font = Font(bold=True, name="Arial", size=10, color="1F4E79")
        ws3.merge_cells(start_row=fila3, start_column=1, end_row=fila3, end_column=1)
        ws3.cell(fila3 + 1, 1, texto).font = Font(name="Arial", size=9)
        ws3.cell(fila3 + 1, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws3.merge_cells(start_row=fila3 + 1, start_column=1, end_row=fila3 + 1, end_column=2)
        ws3.row_dimensions[fila3 + 1].height = 60
        fila3 += 3
    ws3.column_dimensions["A"].width = 90

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre_archivo = f"kpi_score_{data['periodo']}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre_archivo}"'}
    )
