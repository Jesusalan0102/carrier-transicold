import io
from datetime import datetime
from zoneinfo import ZoneInfo
TZ = ZoneInfo("America/Tijuana")
import pymysql
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from db import get_db_connection, execute_read
from auth import verify_token
from pydantic import BaseModel

router = APIRouter(prefix="/reportes", tags=["Reportes Maestros"])

# ── Paleta corporativa ────────────────────────────────────────────────────────
AZUL_CORP   = "1F4E78"   # encabezados principales
AZUL_CLARO  = "2E75B6"   # encabezados secundarios
VERDE       = "375623"   # horarios / asistencia
AMARILLO    = "FFF2CC"   # retardo leve
ROJO_CLARO  = "FFDADA"   # retardo severo
GRIS_FILA   = "F2F2F2"   # filas alternas

def _hdr(color=AZUL_CORP):
    return {
        "font":      Font(name="Arial", size=10, bold=True, color="FFFFFF"),
        "fill":      PatternFill("solid", start_color=color, end_color=color),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border":    Border(
            bottom=Side(style="thin", color="FFFFFF"),
            right=Side(style="thin", color="FFFFFF"),
        ),
    }

def _apply(cell, styles):
    for attr, val in styles.items():
        setattr(cell, attr, val)

def _autofit(ws, min_w=10, max_w=45):
    for col in ws.columns:
        length = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)

def _write_sheet(ws, columns, rows, hdr_color=AZUL_CORP, freeze=True):
    """Escribe encabezados + filas con formato alterno."""
    ws.append(columns)
    hstyle = _hdr(hdr_color)
    for col_i in range(1, len(columns) + 1):
        _apply(ws.cell(1, col_i), hstyle)
    ws.row_dimensions[1].height = 22

    for row_i, row in enumerate(rows, 2):
        ws.append(row)
        if row_i % 2 == 0:
            fill = PatternFill("solid", start_color=GRIS_FILA, end_color=GRIS_FILA)
            for col_i in range(1, len(columns) + 1):
                ws.cell(row_i, col_i).fill = fill

    if freeze:
        ws.freeze_panes = "A2"
    _autofit(ws)


def _safe_str(v):
    if isinstance(v, (bytes, bytearray)):
        return "[Archivo Binario]"
    if v is None:
        return ""
    return str(v)


def _query(conn, sql, params=None):
    with conn.cursor(pymysql.cursors.DictCursor) as cur:
        cur.execute(sql, params or ())
        return cur.fetchall()


# ── Hoja 1: Resumen KPIs ──────────────────────────────────────────────────────
def _sheet_resumen(wb, conn):
    ws = wb.create_sheet("Resumen_KPIs")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18

    titulo = ws.cell(1, 1, "REPORTE MAESTRO — CARRIER TRANSICOLD")
    titulo.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    titulo.fill = PatternFill("solid", start_color=AZUL_CORP, end_color=AZUL_CORP)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 28

    fecha = ws.cell(2, 1, f"Generado: {datetime.now(TZ).strftime('%d/%m/%Y %H:%M')}")
    fecha.font = Font(name="Arial", size=9, italic=True, color="595959")
    ws.merge_cells("A2:B2")

    ws.append([])

    kpis = []
    try:
        kpis.append(("Unidades Registradas",   _query(conn, "SELECT COUNT(*) c FROM unidades WHERE oculto=0")[0]["c"]))
        kpis.append(("Tickets Abiertos",        _query(conn, "SELECT COUNT(*) c FROM tickets WHERE atendido=0")[0]["c"]))
        kpis.append(("Tickets Cerrados",        _query(conn, "SELECT COUNT(*) c FROM tickets WHERE atendido=1")[0]["c"]))
        kpis.append(("Asignaciones Totales",    _query(conn, "SELECT COUNT(*) c FROM asignaciones")[0]["c"]))
        kpis.append(("Técnicos Activos",        _query(conn, "SELECT COUNT(*) c FROM users WHERE role != 'inactivo'")[0]["c"]))
        kpis.append(("Registros Asistencia",    _query(conn, "SELECT COUNT(*) c FROM registros_asistencia")[0]["c"]))
        kpis.append(("Horarios Configurados",   _query(conn, "SELECT COUNT(*) c FROM horarios")[0]["c"]))
    except Exception:
        pass

    label_style = Font(name="Arial", size=10, bold=True)
    val_style   = Font(name="Arial", size=11, bold=True, color="1F4E78")

    for row_i, (label, val) in enumerate(kpis, 4):
        ws.cell(row_i, 1, label).font = label_style
        c = ws.cell(row_i, 2, val)
        c.font = val_style
        c.alignment = Alignment(horizontal="center")
        if row_i % 2 == 0:
            for col_i in (1, 2):
                ws.cell(row_i, col_i).fill = PatternFill("solid", start_color=GRIS_FILA, end_color=GRIS_FILA)


# ── Hoja 2: Series / Unidades ─────────────────────────────────────────────────
def _sheet_unidades(wb, conn):
    ws = wb.create_sheet("Series_Unidades")
    rows_db = _query(conn, """
        SELECT
            unit_number              AS `Número de Unidad`,
            id_lote                  AS `Lote / Flota`,
            vin_number               AS `VIN (Chasis)`,
            reefer_model             AS `Modelo Reefer`,
            reefer_serial            AS `Serie Reefer`,
            evaporator_model_1       AS `Modelo Evaporador 1`,
            evaporator_serial_mjs11  AS `Serie Evaporador 1`,
            evaporator_model_2       AS `Modelo Evaporador 2`,
            evaporator_serial_mjd22  AS `Serie Evaporador 2`,
            engine_serial            AS `Serie Motor`,
            compressor_serial        AS `Serie Compresor`,
            generator_serial         AS `Serie Generador`,
            battery_charger_serial   AS `Serie Cargador Batería`,
            fecha_registro           AS `Fecha y Hora de Registro`
        FROM unidades
        WHERE oculto = 0
        ORDER BY id_lote, unit_number
    """)
    if not rows_db:
        ws.cell(1, 1, "Sin registros").font = Font(italic=True)
        return

    cols = list(rows_db[0].keys())

    # Encabezado
    ws.append(cols)
    hstyle = _hdr(AZUL_CORP)
    for col_i in range(1, len(cols) + 1):
        _apply(ws.cell(1, col_i), hstyle)
    ws.row_dimensions[1].height = 22

    fill_alt = PatternFill("solid", start_color=GRIS_FILA, end_color=GRIS_FILA)

    for row_i, r in enumerate(rows_db, 2):
        row_vals = []
        for c in cols:
            val = r[c]
            # Formatear fecha/hora en Tijuana
            if c == "Fecha y Hora de Registro":
                if val is None:
                    row_vals.append("—")
                elif hasattr(val, "strftime"):
                    # Si viene como datetime naive desde DB, mostrarlo directo
                    row_vals.append(val.strftime("%d/%m/%Y %H:%M:%S"))
                else:
                    row_vals.append(_safe_str(val))
            else:
                row_vals.append(_safe_str(val) if val not in (None, "") else "—")
        ws.append(row_vals)

        if row_i % 2 == 0:
            for col_i in range(1, len(cols) + 1):
                ws.cell(row_i, col_i).fill = fill_alt

    ws.freeze_panes = "A2"
    _autofit(ws)


# ── Hoja 3: Actividades ───────────────────────────────────────────────────────
def _sheet_actividades(wb, conn):
    ws = wb.create_sheet("Actividades")
    rows_db = _query(conn, """
        SELECT a.id, a.unidad, a.actividad_id, a.tecnico, a.estado,
               COALESCE(c.comentarios, a.comentario) AS comentario,
               a.fecha_asignacion, a.fecha_inicio, a.fecha_fin, a.ticket_id
        FROM asignaciones a
        INNER JOIN unidades u ON u.unit_number = a.unidad
        LEFT JOIN (
            SELECT asignacion_id,
                   GROUP_CONCAT(
                       CONCAT(tecnico, ' (', DATE_FORMAT(fecha, '%%d/%%m/%%Y %%H:%%i'), '): ', comentario)
                       ORDER BY fecha SEPARATOR '  |  '
                   ) AS comentarios
            FROM comentarios_actividades
            GROUP BY asignacion_id
        ) c ON c.asignacion_id = a.id
        WHERE u.oculto = 0
        ORDER BY a.fecha_asignacion DESC
    """)
    if not rows_db:
        ws.cell(1, 1, "Sin registros").font = Font(italic=True)
        return
    cols = list(rows_db[0].keys())
    _write_sheet(ws, cols, [[_safe_str(r[c]) for c in cols] for r in rows_db])


# ── Hoja 4: Tickets ───────────────────────────────────────────────────────────
def _sheet_tickets(wb, conn):
    ws = wb.create_sheet("Tickets")
    rows_db = _query(conn, """
        SELECT t.ticket_num AS `Ticket #`,
               t.unit_number AS Unidad,
               u.vin_number AS VIN,
               t.descripcion AS `Descripción / Problema`,
               t.creado_por AS `Creado Por`,
               a.tecnico AS `Técnico Asignado`,
               t.atendido AS Atendido,
               t.reporte_enviado AS `Reporte Enviado`,
               COALESCE(t.reporte_texto, '—') AS `Reporte Final del Técnico`,
               t.fecha_atencion AS `Fecha Atención`,
               t.fecha_reporte AS `Fecha Reporte`
        FROM tickets t
        LEFT JOIN unidades u ON u.unit_number = t.unit_number
        LEFT JOIN asignaciones a ON a.ticket_id = t.id
        WHERE COALESCE(u.oculto, 0) = 0
        ORDER BY t.ticket_num DESC
    """)
    if not rows_db:
        ws.cell(1, 1, "Sin registros").font = Font(italic=True)
        return
    cols = list(rows_db[0].keys())
    _write_sheet(ws, cols, [[_safe_str(r[c]) for c in cols] for r in rows_db])


# ── Hoja 6: Horarios Semanales ────────────────────────────────────────────────
def _sheet_horarios(wb, conn):
    ws = wb.create_sheet("Horarios_Semanales")

    # Pivot: técnico × día
    rows_db = _query(conn, """
        SELECT username, fecha, semana, hora_entrada, hora_salida
        FROM horarios
        ORDER BY username, fecha
    """)

    cols = ["Técnico (username)", "Nombre Completo",
            "Lunes Entrada",  "Lunes Salida",
            "Martes Entrada", "Martes Salida",
            "Miércoles Entrada", "Miércoles Salida",
            "Jueves Entrada", "Jueves Salida",
            "Viernes Entrada","Viernes Salida",
            "Sábado Entrada", "Sábado Salida"]

    _write_sheet(ws, cols, [], hdr_color=VERDE)

    # Obtener nombres completos de users si hay columna nombre
    try:
        users_db = _query(conn, "SELECT username, username AS nombre FROM users")
        # try nombre_completo column
        try:
            users_db = _query(conn, "SELECT username, nombre_completo AS nombre FROM users")
        except Exception:
            pass
        nombres = {u["username"]: u["nombre"] for u in (users_db or [])}
    except Exception:
        nombres = {}

    # Agrupar por técnico
    from collections import defaultdict
    import datetime as _dt

    tec_data = defaultdict(dict)  # {username: {weekday_name: {entrada, salida}}}
    DAYS = {0: "Lunes", 1: "Martes", 2: "Miércoles", 3: "Jueves", 4: "Viernes", 5: "Sábado"}

    for r in (rows_db or []):
        try:
            if hasattr(r["fecha"], "weekday"):
                wd = r["fecha"].weekday()
            else:
                wd = _dt.date.fromisoformat(str(r["fecha"])).weekday()
            day_name = DAYS.get(wd)
            if day_name:
                def _fmt_time(t):
                    if t is None:
                        return ""
                    if hasattr(t, "seconds"):  # timedelta
                        h, rem = divmod(t.seconds, 3600)
                        return f"{h:02d}:{rem//60:02d}"
                    return str(t)[:5]
                tec_data[r["username"]][day_name + " Entrada"] = _fmt_time(r["hora_entrada"])
                tec_data[r["username"]][day_name + " Salida"]  = _fmt_time(r["hora_salida"])
        except Exception:
            continue

    row_i = 2
    for tec, days in sorted(tec_data.items()):
        row = [
            tec,
            nombres.get(tec, tec),
            days.get("Lunes Entrada", ""),    days.get("Lunes Salida", ""),
            days.get("Martes Entrada", ""),   days.get("Martes Salida", ""),
            days.get("Miércoles Entrada", ""),days.get("Miércoles Salida", ""),
            days.get("Jueves Entrada", ""),   days.get("Jueves Salida", ""),
            days.get("Viernes Entrada", ""),  days.get("Viernes Salida", ""),
            days.get("Sábado Entrada", ""),   days.get("Sábado Salida", ""),
        ]
        ws.append(row)
        if row_i % 2 == 0:
            for ci in range(1, len(cols) + 1):
                ws.cell(row_i, ci).fill = PatternFill("solid", start_color=GRIS_FILA, end_color=GRIS_FILA)
        row_i += 1

    ws.freeze_panes = "A2"
    _autofit(ws)


# ── Hoja 7: Asistencia y Retardos ─────────────────────────────────────────────
def _sheet_asistencia(wb, conn):
    ws = wb.create_sheet("Asistencia_Retardos")

    cols = ["Fecha", "Técnico", "Tipo", "Hora Checkin",
            "Horario Programado Entrada", "Horario Programado Salida",
            "Retardo (min)", "Distancia (m)", "Aprobado", "Estado"]

    _write_sheet(ws, cols, [], hdr_color=VERDE)

    # Registros reales
    regs = _query(conn, """
        SELECT r.fecha, r.username, r.tipo, r.hora_checkin,
               r.distancia_metros, r.aprobado,
               COALESCE(r.retardo_min, 0) AS retardo_min
        FROM registros_asistencia r
        ORDER BY r.fecha DESC, r.hora_checkin
    """) or []

    # Horarios programados index por (username, fecha)
    horarios_db = _query(conn, "SELECT username, fecha, hora_entrada, hora_salida FROM horarios") or []
    from collections import defaultdict
    import datetime as _dt

    def _fmt_t(t):
        if t is None: return ""
        if hasattr(t, "seconds"):
            h, rem = divmod(t.seconds, 3600)
            return f"{h:02d}:{rem//60:02d}"
        return str(t)[:5]

    hor_idx = {}
    for h in horarios_db:
        try:
            if hasattr(h["fecha"], "weekday"):
                wd = h["fecha"].weekday()
                fecha_str = h["fecha"].isoformat()
            else:
                d = _dt.date.fromisoformat(str(h["fecha"]))
                wd = d.weekday()
                fecha_str = str(h["fecha"])
            hor_idx[(h["username"], fecha_str)] = (
                _fmt_t(h["hora_entrada"]),
                _fmt_t(h["hora_salida"])
            )
        except Exception:
            continue

    fill_amarillo = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
    fill_rojo     = PatternFill("solid", start_color="FFDADA", end_color="FFDADA")
    fill_verde    = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")

    row_i = 2
    for r in regs:
        try:
            fecha_str = r["fecha"].isoformat() if hasattr(r["fecha"], "isoformat") else str(r["fecha"])
        except Exception:
            fecha_str = str(r.get("fecha", ""))

        hora_checkin = str(r.get("hora_checkin", "") or "")[:5]
        retardo      = int(r.get("retardo_min") or 0)
        aprobado     = r.get("aprobado", 0)
        distancia    = r.get("distancia_metros", "")
        tipo         = r.get("tipo", "")

        hor_e, hor_s = hor_idx.get((r["username"], fecha_str), ("", ""))

        if not aprobado:
            estado = "❌ Fuera de geocerca"
        elif retardo >= 30:
            estado = f"⚠️ Retardo severo +{retardo}min"
        elif retardo > 0:
            estado = f"⏱ Retardo +{retardo}min"
        else:
            estado = "✅ A tiempo"

        row = [fecha_str, r["username"], tipo, hora_checkin,
               hor_e, hor_s,
               retardo if retardo else "",
               distancia, "Sí" if aprobado else "No", estado]
        ws.append(row)

        # Color de fila según estado
        if retardo >= 30:
            fill = fill_rojo
        elif retardo > 0:
            fill = fill_amarillo
        elif aprobado:
            fill = fill_verde
        else:
            fill = fill_rojo

        for ci in range(1, len(cols) + 1):
            ws.cell(row_i, ci).fill = fill

        row_i += 1

    ws.freeze_panes = "A2"
    _autofit(ws)


# ── Hoja 8: Resumen Retardos por Técnico ─────────────────────────────────────
def _sheet_resumen_retardos(wb, conn):
    ws = wb.create_sheet("Resumen_Retardos_Tecnico")

    cols = ["Técnico", "Total Entradas", "Entradas a Tiempo",
            "Con Retardo", "Retardo Promedio (min)", "Retardo Máx (min)",
            "Registros Fuera Geocerca", "Días con Asistencia"]

    _write_sheet(ws, cols, [], hdr_color=AZUL_CLARO)

    rows_db = _query(conn, """
        SELECT
            username,
            SUM(tipo = 'entrada') AS total_entradas,
            SUM(tipo = 'entrada' AND COALESCE(retardo_min,0) = 0) AS a_tiempo,
            SUM(tipo = 'entrada' AND COALESCE(retardo_min,0) > 0) AS con_retardo,
            ROUND(AVG(CASE WHEN tipo='entrada' THEN COALESCE(retardo_min,0) END), 1) AS retardo_prom,
            MAX(CASE WHEN tipo='entrada' THEN COALESCE(retardo_min,0) END) AS retardo_max,
            SUM(aprobado = 0) AS fuera_geocerca,
            COUNT(DISTINCT fecha) AS dias_asistencia
        FROM registros_asistencia
        GROUP BY username
        ORDER BY con_retardo DESC, username
    """) or []

    row_i = 2
    fill_warn = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")

    for r in rows_db:
        con_ret = int(r.get("con_retardo") or 0)
        row = [
            r["username"],
            r.get("total_entradas", 0),
            r.get("a_tiempo", 0),
            con_ret,
            r.get("retardo_prom", 0),
            r.get("retardo_max", 0),
            r.get("fuera_geocerca", 0),
            r.get("dias_asistencia", 0),
        ]
        ws.append(row)
        if con_ret > 3:
            for ci in range(1, len(cols) + 1):
                ws.cell(row_i, ci).fill = fill_warn
        row_i += 1

    ws.freeze_panes = "A2"
    _autofit(ws)


# ── Hoja 9: Inventario ────────────────────────────────────────────────────────
def _sheet_inventario(wb, conn):
    ws = wb.create_sheet("Inventario")
    try:
        rows_db = _query(conn, "SELECT * FROM inventario_data ORDER BY id")
    except Exception:
        try:
            rows_db = _query(conn, "SELECT * FROM inventario ORDER BY id")
        except Exception:
            ws.cell(1, 1, "Tabla de inventario no encontrada").font = Font(italic=True)
            return
    if not rows_db:
        ws.cell(1, 1, "Sin registros").font = Font(italic=True)
        return
    cols = list(rows_db[0].keys())
    _write_sheet(ws, cols, [[_safe_str(r[c]) for c in cols] for r in rows_db])


# ── Hoja 10: Usuarios ─────────────────────────────────────────────────────────
def _sheet_usuarios(wb, conn):
    ws = wb.create_sheet("Usuarios")
    rows_db = _query(conn, """
        SELECT id, username, role,
               CASE WHEN role = 'inactivo' THEN 'No' ELSE 'Sí' END AS activo
        FROM users ORDER BY role, username
    """) or []
    cols = ["ID", "Username", "Rol", "Activo"]
    _write_sheet(ws, cols, [[_safe_str(r[c]) for c in list(r.keys())] for r in rows_db])


# ═════════════════════════════════════════════════════════════════════════
# ── REPORTE POR LOTE ────────────────────────────────────────────────────
# ═════════════════════════════════════════════════════════════════════════

def _sheet_lote_resumen(wb, conn, id_lote):
    ws = wb.create_sheet("Resumen_Lote")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20

    titulo = ws.cell(1, 1, f"REPORTE DE LOTE — {id_lote}")
    titulo.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    titulo.fill = PatternFill("solid", start_color=AZUL_CORP, end_color=AZUL_CORP)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 28

    fecha = ws.cell(2, 1, f"Generado: {datetime.now(TZ).strftime('%d/%m/%Y %H:%M')}")
    fecha.font = Font(name="Arial", size=9, italic=True, color="595959")
    ws.merge_cells("A2:B2")
    ws.append([])

    total_u = _query(conn, "SELECT COUNT(*) c FROM unidades WHERE id_lote=%s", (id_lote,))[0]["c"]
    unit_numbers_rows = _query(conn, "SELECT unit_number FROM unidades WHERE id_lote=%s", (id_lote,))
    unit_numbers = [r["unit_number"] for r in unit_numbers_rows]

    if unit_numbers:
        placeholders = ",".join(["%s"] * len(unit_numbers))
        asigs = _query(conn, f"SELECT estado FROM asignaciones WHERE unidad IN ({placeholders})", tuple(unit_numbers)) or []
        tickets = _query(conn, f"SELECT atendido FROM tickets WHERE unit_number IN ({placeholders})", tuple(unit_numbers)) or []
    else:
        asigs, tickets = [], []

    completadas = sum(1 for a in asigs if a["estado"] == "completada")
    en_proceso  = sum(1 for a in asigs if a["estado"] == "en_proceso")
    pendientes  = sum(1 for a in asigs if a["estado"] == "pendiente")
    total_t = len(asigs)
    avance_pct = round(completadas / total_t * 100, 1) if total_t else 0
    tk_total = len(tickets)
    tk_abiertos = sum(1 for t in tickets if not t["atendido"])

    kpis = [
        ("Unidades en el Lote",       total_u),
        ("Actividades Completadas",   completadas),
        ("Actividades En Proceso",    en_proceso),
        ("Actividades Pendientes",    pendientes),
        ("% Avance del Lote",         f"{avance_pct}%"),
        ("Tickets Totales",           tk_total),
        ("Tickets Abiertos",          tk_abiertos),
    ]
    label_style = Font(name="Arial", size=10, bold=True)
    val_style   = Font(name="Arial", size=11, bold=True, color="1F4E78")
    for row_i, (label, val) in enumerate(kpis, 4):
        ws.cell(row_i, 1, label).font = label_style
        c = ws.cell(row_i, 2, val)
        c.font = val_style
        c.alignment = Alignment(horizontal="center")
        if row_i % 2 == 0:
            for col_i in (1, 2):
                ws.cell(row_i, col_i).fill = PatternFill("solid", start_color=GRIS_FILA, end_color=GRIS_FILA)


def _sheet_lote_unidades(wb, conn, id_lote):
    ws = wb.create_sheet("Series_Unidades")
    rows_db = _query(conn, """
        SELECT
            unit_number              AS `Número de Unidad`,
            vin_number               AS `VIN (Chasis)`,
            reefer_model             AS `Modelo Reefer`,
            reefer_serial            AS `Serie Reefer`,
            evaporator_model_1       AS `Modelo Evaporador 1`,
            evaporator_serial_mjs11  AS `Serie Evaporador 1`,
            evaporator_model_2       AS `Modelo Evaporador 2`,
            evaporator_serial_mjd22  AS `Serie Evaporador 2`,
            engine_serial            AS `Serie Motor`,
            compressor_serial        AS `Serie Compresor`,
            generator_serial         AS `Serie Generador`,
            battery_charger_serial   AS `Serie Cargador Batería`,
            fecha_registro           AS `Fecha y Hora de Registro`,
            oculto                   AS `Oculto`
        FROM unidades
        WHERE id_lote=%s
        ORDER BY unit_number
    """, (id_lote,))
    if not rows_db:
        ws.cell(1, 1, "Sin unidades en este lote").font = Font(italic=True)
        return
    cols = list(rows_db[0].keys())
    rows_fmt = []
    for r in rows_db:
        row_vals = []
        for c in cols:
            val = r[c]
            if c == "Fecha y Hora de Registro" and hasattr(val, "strftime"):
                row_vals.append(val.strftime("%d/%m/%Y %H:%M:%S"))
            elif c == "Oculto":
                row_vals.append("Sí" if val else "No")
            else:
                row_vals.append(_safe_str(val) if val not in (None, "") else "—")
        rows_fmt.append(row_vals)
    _write_sheet(ws, cols, rows_fmt)


def _sheet_lote_actividades(wb, conn, id_lote):
    ws = wb.create_sheet("Actividades")
    rows_db = _query(conn, """
        SELECT a.unidad AS Unidad, a.actividad_id AS Actividad, a.tecnico AS Técnico,
               a.estado AS Estado, COALESCE(c.comentarios, a.comentario) AS Comentario,
               a.fecha_asignacion AS `Fecha Asignación`,
               a.fecha_inicio AS `Fecha Inicio`, a.fecha_fin AS `Fecha Fin`
        FROM asignaciones a
        INNER JOIN unidades u ON u.unit_number = a.unidad
        LEFT JOIN (
            SELECT asignacion_id,
                   GROUP_CONCAT(
                       CONCAT(tecnico, ' (', DATE_FORMAT(fecha, '%%d/%%m/%%Y %%H:%%i'), '): ', comentario)
                       ORDER BY fecha SEPARATOR '  |  '
                   ) AS comentarios
            FROM comentarios_actividades
            GROUP BY asignacion_id
        ) c ON c.asignacion_id = a.id
        WHERE u.id_lote=%s
        ORDER BY a.unidad, a.fecha_asignacion DESC
    """, (id_lote,))
    if not rows_db:
        ws.cell(1, 1, "Sin actividades registradas para este lote").font = Font(italic=True)
        return
    cols = list(rows_db[0].keys())
    _write_sheet(ws, cols, [[_safe_str(r[c]) for c in cols] for r in rows_db])
    colorear_idx = cols.index("Estado") + 1 if "Estado" in cols else None
    if colorear_idx:
        COLOR_MAP = {
            "completada": ("C6EFCE", "276221"),
            "en_proceso": ("DDEBF7", "1F4E79"),
            "pendiente":  ("FFEB9C", "7D6608"),
            "solicitado": ("FFEB9C", "7D6608"),
        }
        for row_i in range(2, len(rows_db) + 2):
            cell = ws.cell(row_i, colorear_idx)
            colors = COLOR_MAP.get(str(cell.value or "").lower())
            if colors:
                cell.fill = PatternFill("solid", start_color=colors[0], end_color=colors[0])
                cell.font = Font(name="Arial", size=9, color=colors[1], bold=True)


def _sheet_lote_tickets(wb, conn, id_lote):
    ws = wb.create_sheet("Tickets")
    rows_db = _query(conn, """
        SELECT t.ticket_num AS `Ticket #`,
               t.unit_number AS Unidad,
               t.descripcion AS `Descripción / Problema`,
               t.creado_por AS `Creado Por`,
               t.atendido AS Atendido,
               t.reporte_enviado AS `Reporte Enviado`,
               COALESCE(t.reporte_texto, '—') AS `Reporte Final del Técnico`,
               t.fecha_creacion AS `Fecha Creación`,
               t.fecha_atencion AS `Fecha Atención`
        FROM tickets t
        INNER JOIN unidades u ON u.unit_number = t.unit_number
        WHERE u.id_lote=%s
        ORDER BY t.ticket_num DESC
    """, (id_lote,))
    if not rows_db:
        ws.cell(1, 1, "Sin tickets registrados para este lote").font = Font(italic=True)
        return
    cols = list(rows_db[0].keys())
    _write_sheet(ws, cols, [[_safe_str(r[c]) for c in cols] for r in rows_db])


@router.get("/lote")
def exportar_reporte_lote(
    id_lote: str = Query(..., description="ID del lote"),
    current_user=Depends(verify_token)
):
    """
    Exporta un reporte Excel enfocado únicamente en las unidades de un lote:
    resumen de avance, series, actividades y tickets de ese lote.
    """
    connection = get_db_connection()
    if not connection:
        raise HTTPException(status_code=500, detail="No hay conexión con la base de datos")

    try:
        existe = _query(connection, "SELECT COUNT(*) c FROM unidades WHERE id_lote=%s", (id_lote,))
        if not existe or existe[0]["c"] == 0:
            raise HTTPException(status_code=404, detail=f"No se encontró el lote '{id_lote}'")

        wb = Workbook()
        wb.remove(wb.active)

        builders = [
            ("Resumen_Lote",    lambda wb, conn: _sheet_lote_resumen(wb, conn, id_lote)),
            ("Series_Unidades", lambda wb, conn: _sheet_lote_unidades(wb, conn, id_lote)),
            ("Actividades",     lambda wb, conn: _sheet_lote_actividades(wb, conn, id_lote)),
            ("Tickets",         lambda wb, conn: _sheet_lote_tickets(wb, conn, id_lote)),
        ]

        for name, fn in builders:
            try:
                fn(wb, connection)
            except Exception as e:
                ws = wb.create_sheet(name)
                ws.cell(1, 1, f"Error al generar esta hoja: {e}").font = Font(italic=True, color="FF0000")
                print(f"[reporte_lote] Hoja '{name}' falló: {e}")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        fecha_hoy = datetime.now(TZ).strftime("%Y%m%d_%H%M")
        nombre_seguro = "".join(c for c in id_lote if c.isalnum() or c in ("-", "_")) or "lote"
        filename = f"Reporte_Lote_{nombre_seguro}_{fecha_hoy}.xlsx"

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error al generar reporte de lote '{id_lote}': {e}")
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}")
    finally:
        connection.close()


class UnidadesSeleccionadas(BaseModel):
    unidades: list[str]


@router.post("/lotes-seleccionados")
def exportar_reporte_unidades_seleccionadas(
    data: UnidadesSeleccionadas,
    current_user=Depends(verify_token)
):
    """
    Genera un único Excel con la hoja Series_Unidades (mismas columnas que el
    reporte por lote) pero filtrado a una lista explícita de unit_numbers,
    posiblemente de varios lotes distintos. Usado desde el Schedule al
    seleccionar filas y elegir qué unidades incluir en el reporte.
    """
    if not data.unidades:
        raise HTTPException(status_code=400, detail="Debes indicar al menos una unidad")

    connection = get_db_connection()
    if not connection:
        raise HTTPException(status_code=500, detail="No hay conexión con la base de datos")

    try:
        placeholders = ",".join(["%s"] * len(data.unidades))
        rows_db = _query(connection, f"""
            SELECT
                id_lote                  AS `Lote`,
                unit_number               AS `Número de Unidad`,
                vin_number               AS `VIN (Chasis)`,
                reefer_model             AS `Modelo Reefer`,
                reefer_serial            AS `Serie Reefer`,
                evaporator_model_1       AS `Modelo Evaporador 1`,
                evaporator_serial_mjs11  AS `Serie Evaporador 1`,
                evaporator_model_2       AS `Modelo Evaporador 2`,
                evaporator_serial_mjd22  AS `Serie Evaporador 2`,
                engine_serial            AS `Serie Motor`,
                compressor_serial        AS `Serie Compresor`,
                generator_serial         AS `Serie Generador`,
                battery_charger_serial   AS `Serie Cargador Batería`,
                fecha_registro           AS `Fecha y Hora de Registro`
            FROM unidades
            WHERE unit_number IN ({placeholders})
            ORDER BY id_lote, unit_number
        """, tuple(data.unidades))

        if not rows_db:
            raise HTTPException(status_code=404, detail="No se encontraron unidades para esa selección")

        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("Series_Unidades")

        cols = list(rows_db[0].keys())
        rows_fmt = []
        for r in rows_db:
            row_vals = []
            for c in cols:
                val = r[c]
                if c == "Fecha y Hora de Registro" and hasattr(val, "strftime"):
                    row_vals.append(val.strftime("%d/%m/%Y %H:%M:%S"))
                else:
                    row_vals.append(_safe_str(val) if val not in (None, "") else "—")
            rows_fmt.append(row_vals)
        _write_sheet(ws, cols, rows_fmt)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        fecha_hoy = datetime.now(TZ).strftime("%Y%m%d_%H%M")
        filename = f"Reporte_Series_Seleccion_{fecha_hoy}.xlsx"

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error al generar reporte de unidades seleccionadas: {e}")
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}")
    finally:
        connection.close()


# ── Endpoint principal ────────────────────────────────────────────────────────
@router.get("/exportar-maestro")
def exportar_sistema_completo(current_user=Depends(verify_token)):
    """
    Exporta el sistema completo a Excel con formato profesional.
    Incluye: KPIs, Unidades, Actividades, Tickets, Cierre, Horarios, Asistencia, Retardos.
    """
    connection = get_db_connection()
    if not connection:
        raise HTTPException(status_code=500, detail="No hay conexión con la base de datos")

    try:
        wb = Workbook()
        wb.remove(wb.active)   # eliminar hoja en blanco por defecto

        builders = [
            ("Resumen_KPIs",             _sheet_resumen),
            ("Series_Unidades",          _sheet_unidades),
            ("Actividades",              _sheet_actividades),
            ("Tickets",                  _sheet_tickets),
            ("Horarios_Semanales",       _sheet_horarios),
            ("Asistencia_Retardos",      _sheet_asistencia),
            ("Resumen_Retardos_Tecnico", _sheet_resumen_retardos),
            ("Inventario",               _sheet_inventario),
            ("Usuarios",                 _sheet_usuarios),
        ]

        for name, fn in builders:
            try:
                fn(wb, connection)
            except Exception as e:
                # Si una hoja falla, la añadimos con nota de error en lugar de abortar todo
                ws = wb.create_sheet(name)
                ws.cell(1, 1, f"Error al generar esta hoja: {e}").font = Font(italic=True, color="FF0000")
                print(f"[reporte] Hoja '{name}' falló: {e}")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        fecha_hoy = datetime.now(TZ).strftime("%Y%m%d_%H%M")
        filename  = f"Reporte_Maestro_CarrierTransicold_{fecha_hoy}.xlsx"

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        print(f"Error al generar reporte maestro: {e}")
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}")

    finally:
        connection.close()
