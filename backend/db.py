import io
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Importamos la función de lectura limpia desde tu db.py
from db import execute_read 
from auth import verify_token

router = APIRouter(
    prefix="/reportes",
    tags=["Reportes Maestros"]
)

@router.get("/exportar-maestro")
def exportar_sistema_completo(current_user=Depends(verify_token)):
    """
    Genera y descarga un archivo Excel (.xlsx) que contiene absolutamente
    toda la información del sistema de Carrier Transicold, organizando
    cada tabla de la base de datos TiDB en una pestaña independiente.
    """
    try:
        # 1. Inicializar el libro de Excel en blanco
        wb = Workbook()
        # Remover la primera hoja que openpyxl crea por defecto para personalizar el orden
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # 2. Lista oficial de las 22 tablas que administra el sistema de Carrier Transicold
        tablas_sistema = [
            "actividades", "asignaciones", "asistencia", "asistencia_config", 
            "asistencia_registros", "comentarios_actividades", "config_sistema", 
            "configuracion_geocerca", "evidencias", "horarios", "inventario", 
            "inventario_columnas", "inventario_data", "inventarios", "lotes", 
            "registros_asistencia", "tickets", "toma_valores_campos", 
            "toma_valores_datos", "unidades", "users", "valores_registrados"
        ]
        
        # 3. Estilos de diseño corporativo (Encabezado Azul Marino elegante)
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        alignment_center = Alignment(horizontal="center", vertical="center")
        
        # 4. Iterar y procesar cada tabla del sistema de manera dinámica
        for nombre_tabla in tablas_sistema:
            # Creamos la query de manera segura usando backticks para proteger nombres reservados
            sql = f"SELECT * FROM `{nombre_tabla}`"
            
            # Ejecutamos la consulta usando la función de lectura de tu db.py (retorna diccionarios)
            registros = execute_read(sql)
            
            # Crear una pestaña exclusiva para la tabla actual (Truncado a 31 caracteres, límite de Excel)
            ws = wb.create_sheet(title=nombre_tabla[:31])
            
            if registros:
                # Extraemos el nombre de las columnas basándonos en el primer registro
                columnas = list(registros[0].keys())
                
                # Insertar la primera fila con los encabezados
                ws.append(columnas)
                
                # Aplicar estilos estilizados a toda la fila 1 (Encabezados)
                for col_num in range(1, len(columnas) + 1):
                    celda = ws.cell(row=1, column=col_num)
                    celda.font = font_header
                    celda.fill = fill_header
                    celda.alignment = alignment_center
                
                # Recorrer cada registro de la tabla e insertar sus filas de datos
                for reg in registros:
                    fila_datos = []
                    for col in columnas:
                        valor = reg[col]
                        
                        # Control de seguridad contra datos binarios pesados (Blobs, fotos, firmas)
                        if isinstance(valor, (bytes, bytearray)):
                            valor = "[Archivo Binario / Imagen]"
                        elif valor is not None:
                            valor = str(valor)
                            
                        fila_datos.append(valor)
                    ws.append(fila_datos)
            else:
                # Si una tabla está completamente vacía, ponemos un mensaje elegante en lugar de dejarla desierta
                ws.cell(row=1, column=1, value="Sin registros actualmente en esta tabla.").font = Font(italic=True)
            
            # 5. Auto-ajuste inteligente del ancho de las columnas de acuerdo con el texto más largo
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                # Añadir un margen de seguridad de +3 caracteres
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
        # 6. Guardar la estructura completa en la memoria RAM (búfer binario BytesIO)
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Nombre por defecto del reporte al ser descargado
        nombre_archivo = "Reporte_Maestro_Carrier_Transicold.xlsx"
        
        # Retornamos el stream directamente al navegador del administrador
        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"}
        )

    except Exception as e:
        print(f"Error crítico en exportar_sistema_completo: {e}")
        raise HTTPException(status_code=500, detail=f"Error interno del servidor: {str(e)}")
