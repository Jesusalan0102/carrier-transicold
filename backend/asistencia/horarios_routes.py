"""
horarios_routes.py — Rutas de horarios semanales
Fixes aplicados:
  - /resumen: query unificada que cruza horarios + registros_asistencia
    correctamente por username Y fecha (sin desfase de timezone)
  - Inasistencias: técnicos con horario pero SIN ningún registro ahora
    aparecen como "ausente" en vez de desaparecer del resumen
  - Retardos: se recalculan aquí también para que el resumen sea
    consistente con lo que guarda routes.py al momento del check-in
  - /mios: ahora devuelve registros REALES de check-in (no horarios programados)
    para que el historial del técnico muestre sus entradas/salidas reales
"""

from datetime import date, timedelta
from typing import List, Optional

from fastapi import APIRouter, Depends, Query, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from auth import verify_token
from db import execute_read, execute_write

import io
import unicodedata
import traceback

import zoneinfo
from datetime import datetime

TZ_TJ = zoneinfo.ZoneInfo("America/Tijuana")


def ahora_tijuana():
    return datetime.now(TZ_TJ)


router = APIRouter(prefix="/api/horarios", tags=["horarios"])


# ── Alertas persistentes de horario ─────────────────────────────────────────
# Complementa el push (VAPID): el push requiere permiso del navegador y
# muchos técnicos nunca lo aceptan. Esta tabla guarda un aviso "in-app" que
# el técnico ve la próxima vez que abre la app, hasta que lo cierre.

def _ensure_notif_table():
    execute_write("""
        CREATE TABLE IF NOT EXISTS horario_notificaciones (
            id          INT AUTO_INCREMENT PRIMARY KEY,
            username    VARCHAR(100) NOT NULL,
            semana      VARCHAR(20)  NOT NULL,
            visto       TINYINT(1)   NOT NULL DEFAULT 0,
            created_at  DATETIME DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_username_visto (username, visto)
        )
    """)


def registrar_alertas_horario(usernames, semana: str):
    """Crea (o actualiza) el aviso in-app 'tu horario cambió' para cada username.
    Si ya existe un aviso sin leer para esa semana, no duplica — solo refresca
    la fecha para que vuelva a aparecer como reciente."""
    if not usernames:
        return
    _ensure_notif_table()
    for username in usernames:
        existente = execute_read(
            "SELECT id FROM horario_notificaciones WHERE username=%s AND semana=%s AND visto=0",
            (username, semana)
        )
        if existente:
            execute_write(
                "UPDATE horario_notificaciones SET created_at=CURRENT_TIMESTAMP WHERE id=%s",
                (existente[0]["id"],)
            )
        else:
            execute_write(
                "INSERT INTO horario_notificaciones (username, semana) VALUES (%s,%s)",
                (username, semana)
            )


# ── GET /api/horarios/alertas ────────────────────────────────────────────────
# El técnico consulta sus avisos de horario sin leer (banner en su app).

@router.get("/alertas")
def get_mis_alertas(current_user=Depends(verify_token)):
    _ensure_notif_table()
    rows = execute_read(
        "SELECT id, semana, created_at FROM horario_notificaciones "
        "WHERE username=%s AND visto=0 ORDER BY created_at DESC",
        (current_user["username"],)
    ) or []
    return [
        {"id": r["id"], "semana": r["semana"], "creado": _to_str(r["created_at"])}
        for r in rows
    ]


class MarcarVistoPayload(BaseModel):
    ids: List[int]


# ── POST /api/horarios/alertas/marcar-visto ──────────────────────────────────

@router.post("/alertas/marcar-visto")
def marcar_alertas_vistas(payload: MarcarVistoPayload, current_user=Depends(verify_token)):
    if not payload.ids:
        return {"ok": True, "marcados": 0}
    _ensure_notif_table()
    placeholders = ",".join(["%s"] * len(payload.ids))
    execute_write(
        f"UPDATE horario_notificaciones SET visto=1 "
        f"WHERE username=%s AND id IN ({placeholders})",
        (current_user["username"], *payload.ids)
    )
    return {"ok": True, "marcados": len(payload.ids)}


# ── Schemas ────────────────────────────────────────────────────────────────────

class HorarioItem(BaseModel):
    username: str
    fecha: str        # "YYYY-MM-DD"
    semana: str       # "YYYY-MM-DD" (lunes de esa semana)
    hora_entrada: Optional[str] = None   # "HH:MM"
    hora_salida:  Optional[str] = None   # "HH:MM"


class HorariosBatch(BaseModel):
    registros: List[HorarioItem]


# ── Helpers ────────────────────────────────────────────────────────────────────

def _to_str(val) -> Optional[str]:
    """Convierte fecha/timedelta/string a string limpio."""
    if val is None:
        return None
    if hasattr(val, "isoformat"):   # date / datetime
        return val.isoformat()
    if hasattr(val, "seconds"):     # timedelta (TIME de pymysql)
        total = val.seconds
        h = total // 3600
        m = (total % 3600) // 60
        return f"{h:02d}:{m:02d}"
    return str(val)


def _hhmm_to_min(hhmm: Optional[str]) -> Optional[int]:
    if not hhmm:
        return None
    try:
        parts = str(hhmm)[:5].split(":")
        return int(parts[0]) * 60 + int(parts[1])
    except Exception:
        return None


# ── GET /api/horarios/ ─────────────────────────────────────────────────────────

@router.get("/")
def get_horarios(
    semana: Optional[str] = Query(None),
    current_user=Depends(verify_token)
):
    if semana:
        rows = execute_read(
            "SELECT id, username, fecha, semana, hora_entrada, hora_salida "
            "FROM horarios WHERE semana=%s ORDER BY fecha, username",
            (semana,)
        )
    else:
        rows = execute_read(
            "SELECT id, username, fecha, semana, hora_entrada, hora_salida "
            "FROM horarios ORDER BY fecha DESC, username LIMIT 100"
        )

    return [
        {
            "id":           r["id"],
            "username":     r["username"],
            "fecha":        _to_str(r["fecha"]),
            "semana":       _to_str(r["semana"]),
            "hora_entrada": _to_str(r["hora_entrada"]),
            "hora_salida":  _to_str(r["hora_salida"]),
        }
        for r in (rows or [])
    ]


# ── GET /api/horarios/mios ─────────────────────────────────────────────────────
# FIX: ahora devuelve registros REALES de check-in, no horarios programados.
# Así el técnico ve sus entradas/salidas reales en el historial.

@router.get("/mios")
def get_mis_horarios(
    semana: Optional[str] = Query(None),
    current_user=Depends(verify_token)
):
    username = current_user["username"]

    if semana:
        # Calcular fechas de la semana
        try:
            lunes = date.fromisoformat(semana)
        except ValueError:
            raise HTTPException(400, "Formato de semana inválido (esperado YYYY-MM-DD)")
        fechas = [(lunes + timedelta(days=i)).isoformat() for i in range(6)]
        placeholders = ",".join(["%s"] * len(fechas))

        # Horarios programados de la semana
        horarios_rows = execute_read(
            f"SELECT fecha, hora_entrada, hora_salida FROM horarios "
            f"WHERE username=%s AND fecha IN ({placeholders})",
            (username, *fechas)
        ) or []

        # Registros reales de check-in de la semana
        registros_rows = execute_read(
            f"SELECT fecha, tipo, hora_checkin FROM registros_asistencia "
            f"WHERE username=%s AND fecha IN ({placeholders}) "
            f"ORDER BY hora_checkin",
            (username, *fechas)
        ) or []

    else:
        # Últimos 14 días
        horarios_rows = execute_read(
            "SELECT fecha, hora_entrada, hora_salida FROM horarios "
            "WHERE username=%s ORDER BY fecha DESC LIMIT 14",
            (username,)
        ) or []

        registros_rows = execute_read(
            "SELECT fecha, tipo, hora_checkin FROM registros_asistencia "
            "WHERE username=%s ORDER BY fecha DESC, hora_checkin DESC LIMIT 28",
            (username,)
        ) or []

    # Cruzar: para cada fecha, devolver hora programada + hora real
    horarios_map = {}
    for h in horarios_rows:
        f = _to_str(h["fecha"])
        horarios_map[f] = {
            "hora_entrada": _to_str(h["hora_entrada"]),
            "hora_salida":  _to_str(h["hora_salida"]),
        }

    registros_map: dict = {}
    for r in registros_rows:
        f = _to_str(r["fecha"])
        if f not in registros_map:
            registros_map[f] = {"entrada": None, "salida": None}
        tipo = r["tipo"]
        hora = _to_str(r["hora_checkin"])
        if hora and len(hora) >= 5:
            hora = hora[:5]
        if tipo == "entrada" and not registros_map[f]["entrada"]:
            registros_map[f]["entrada"] = hora
        elif tipo == "salida":
            registros_map[f]["salida"] = hora

    # Unión de fechas de ambas fuentes
    todas_fechas = sorted(set(list(horarios_map.keys()) + list(registros_map.keys())), reverse=True)

    resultado = []
    for f in todas_fechas[:14]:
        prog = horarios_map.get(f, {})
        real = registros_map.get(f, {})
        resultado.append({
            "fecha":             f,
            "hora_entrada":      prog.get("hora_entrada"),   # programada
            "hora_salida":       prog.get("hora_salida"),    # programada
            "entrada_real":      real.get("entrada"),        # check-in real
            "salida_real":       real.get("salida"),         # check-out real
        })

    return resultado


# ── POST /api/horarios/ ────────────────────────────────────────────────────────

@router.post("/")
async def save_horarios(payload: HorariosBatch, current_user=Depends(verify_token)):
    """
    Guarda el horario semanal.

    Optimizaciones:
      - 1 sola query para precargar TODOS los horarios existentes del
        username/fecha del payload, en vez de un SELECT por fila (antes eran
        hasta N SELECT + N UPDATE/INSERT secuenciales — esto es lo que hacía
        lento el botón "Guardar Horarios" con varios técnicos).
      - Filas sin cambios reales se saltan por completo (no se escriben en
        la BD), lo que también evita mandar alertas de "horario actualizado"
        quien en realidad no tuvo ningún cambio.

    Al final, si hubo cambios reales, dispara una alerta (WebSocket + push)
    a los técnicos afectados avisándoles que su horario cambió.
    """
    if current_user["role"] not in ("admin",):
        raise HTTPException(status_code=403, detail="Solo administradores")

    if not payload.registros:
        return {"ok": True, "guardados": 0, "eliminados": 0, "notificados": 0}

    usernames_payload = list({item.username for item in payload.registros})
    fechas_payload = list({item.fecha for item in payload.registros})

    existentes = {}
    if usernames_payload and fechas_payload:
        ph_u = ",".join(["%s"] * len(usernames_payload))
        ph_f = ",".join(["%s"] * len(fechas_payload))
        rows = execute_read(
            f"SELECT username, fecha, hora_entrada, hora_salida FROM horarios "
            f"WHERE username IN ({ph_u}) AND fecha IN ({ph_f})",
            (*usernames_payload, *fechas_payload)
        ) or []
        for r in rows:
            existentes[(r["username"], _to_str(r["fecha"]))] = r

    guardados  = 0
    eliminados = 0
    afectados  = set()
    semana_afectada = None

    for item in payload.registros:
        entrada = (item.hora_entrada or "").strip() or None
        salida  = (item.hora_salida  or "").strip() or None

        fila_actual = existentes.get((item.username, item.fecha))
        actual_entrada = _to_str(fila_actual["hora_entrada"]) if fila_actual else None
        actual_salida  = _to_str(fila_actual["hora_salida"])  if fila_actual else None

        if not entrada and not salida:
            if fila_actual:
                execute_write(
                    "DELETE FROM horarios WHERE username=%s AND fecha=%s",
                    (item.username, item.fecha)
                )
                eliminados += 1
                afectados.add(item.username)
                semana_afectada = item.semana
            continue

        # Sin cambios reales respecto a lo guardado: no tocar la BD.
        if fila_actual and entrada == actual_entrada and salida == actual_salida:
            continue

        if fila_actual:
            execute_write(
                "UPDATE horarios SET hora_entrada=%s, hora_salida=%s, semana=%s "
                "WHERE username=%s AND fecha=%s",
                (entrada, salida, item.semana, item.username, item.fecha)
            )
        else:
            execute_write(
                "INSERT INTO horarios (username, fecha, semana, hora_entrada, hora_salida) "
                "VALUES (%s,%s,%s,%s,%s)",
                (item.username, item.fecha, item.semana, entrada, salida)
            )
        guardados += 1
        afectados.add(item.username)
        semana_afectada = item.semana

    if afectados:
        registrar_alertas_horario(sorted(afectados), semana_afectada)
        from routers.ws import notify
        await notify("horario_actualizado", {
            "usernames": sorted(afectados),
            "semana": semana_afectada,
            "cantidad": len(afectados),
        })

    return {
        "ok": True,
        "guardados": guardados,
        "eliminados": eliminados,
        "notificados": len(afectados),
    }


# ── POST /api/horarios/importar-excel ──────────────────────────────────────────
# Parsea un .xlsx con columnas: Técnico (username), Nombre Completo,
# Lunes Entrada, Lunes Salida, ..., Sábado Entrada, Sábado Salida
# Hace fuzzy matching de nombre → username si la columna username está vacía.
# Devuelve un preview JSON para que el admin confirme antes de guardar.

def _normalizar(s: str) -> str:
    """Lowercase, sin tildes, sin espacios extra."""
    s = s.strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return " ".join(s.split())


def _similitud(a: str, b: str) -> float:
    """Ratio de palabras en común (0-1)."""
    wa = set(_normalizar(a).split())
    wb = set(_normalizar(b).split())
    if not wa or not wb:
        return 0.0
    return len(wa & wb) / max(len(wa), len(wb))


DIAS_COLS = [
    ("Lunes",     "lunes_entrada",     "lunes_salida"),
    ("Martes",    "martes_entrada",    "martes_salida"),
    ("Miércoles", "miercoles_entrada", "miercoles_salida"),
    ("Jueves",    "jueves_entrada",    "jueves_salida"),
    ("Viernes",   "viernes_entrada",   "viernes_salida"),
    ("Sábado",    "sabado_entrada",    "sabado_salida"),
]


def _parse_hora(val) -> Optional[str]:
    """Convierte cualquier representacion de hora a HH:MM o None.

    Soporta:
      - Fraccion decimal de dia de Excel  (0.333... -> 08:00)
      - timedelta de pymysql              (timedelta(seconds=28800) -> 08:00)
      - Strings HH:MM o HH:MM:SS
      - "Descansa" / "-" / None           -> None
    """
    if val is None:
        return None

    # timedelta (TIME columns from pymysql)
    import datetime as _dt
    if isinstance(val, _dt.timedelta):
        total = int(val.total_seconds())
        h, remainder = divmod(total, 3600)
        m = remainder // 60
        return f"{h:02d}:{m:02d}"

    # Numero flotante / entero -> fraccion de dia de Excel
    if isinstance(val, (int, float)):
        total_min = round(val * 24 * 60)
        h, m = divmod(total_min, 60)
        if 0 <= h < 24:
            return f"{h:02d}:{m:02d}"
        return None

    s = str(val).strip()
    if not s or s.lower() in ("-", "descansa", "libre", ""):
        return None

    # Formato HH:MM:SS o HH:MM
    parts = s.split(":")
    if len(parts) >= 2:
        try:
            h = int(parts[0])
            m = int(parts[1])
            if 0 <= h < 24 and 0 <= m < 60:
                return f"{h:02d}:{m:02d}"
        except Exception:
            pass

    # Ultimo intento: numero como string ("0.333")
    try:
        frac = float(s)
        total_min = round(frac * 24 * 60)
        h, m = divmod(total_min, 60)
        if 0 <= h < 24:
            return f"{h:02d}:{m:02d}"
    except Exception:
        pass

    return None



@router.post("/importar-excel")
async def importar_excel(
    semana: str = Query(..., description="YYYY-MM-DD (lunes de la semana)"),
    file: UploadFile = File(...),
    current_user=Depends(verify_token)
):
    """
    Recibe un .xlsx, lo parsea y devuelve un preview con el matching
    username → nombre_completo para que el admin confirme.
    """
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores")

    try:
        return await _importar_excel_impl(semana, file)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Error interno: {traceback.format_exc()[-300:]}")


async def _importar_excel_impl(semana: str, file):
    # Validar semana
    try:
        lunes = date.fromisoformat(semana)
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de semana inválido (YYYY-MM-DD)")

    # Calcular fechas de la semana
    fechas_semana = [(lunes + timedelta(days=i)).isoformat() for i in range(6)]

    # Leer archivo
    content = await file.read()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el archivo Excel: {e}")

    # Buscar hoja "Horarios" (o la primera)
    hoja_nombre = "Horarios" if "Horarios" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[hoja_nombre]

    # Leer encabezados de la primera fila
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="El archivo está vacío")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    # Mapear columnas por nombre
    def col_idx(name: str) -> Optional[int]:
        name_n = _normalizar(name)
        for i, h in enumerate(headers):
            if _normalizar(h) == name_n:
                return i
        return None

    # Buscar columna username con nombres alternativos (usar next para evitar problema con índice 0)
    def first_col(*names):
        for name in names:
            idx = col_idx(name)
            if idx is not None:
                return idx
        return None

    idx_username = first_col(
        "Técnico (username)", "Tecnico (username)",
        "Username", "usuario", "user"
    )
    idx_nombre = first_col(
        "Nombre Completo", "Nombre", "nombre_completo", "nombre"
    )
    # Si no hay columna nombre, usar la de username como fallback
    if idx_nombre is None and idx_username is not None:
        idx_nombre = idx_username

    col_map = {}
    for dia_label, key_e, key_s in DIAS_COLS:
        col_map[key_e] = col_idx(f"{dia_label} Entrada")
        col_map[key_s] = col_idx(f"{dia_label} Salida")

    if idx_nombre is None:
        raise HTTPException(status_code=400, detail="No se encontró la columna 'Nombre Completo'")

    # Cargar técnicos de la BD para hacer matching
    # Intentar con nombre_completo primero; si la columna no existe en el schema, caer a username
    try:
        users_db = execute_read(
            "SELECT username, nombre_completo AS nombre FROM users WHERE role IN ('tecnico','lider') AND nombre_completo IS NOT NULL AND nombre_completo != ''"
        ) or []
    except Exception:
        users_db = []

    if not users_db:
        # nombre_completo no existe o vacío para todos -> usar username como referencia
        try:
            users_db = execute_read(
                "SELECT username, username AS nombre FROM users WHERE role IN ('tecnico','lider')"
            ) or []
        except Exception:
            users_db = []

    def buscar_username(nombre_excel: str) -> Optional[str]:
        mejor = None
        mejor_score = 0.0
        for u in users_db:
            score = _similitud(nombre_excel, u["nombre"])
            if score > mejor_score:
                mejor_score = score
                mejor = u["username"]
        return mejor if mejor_score >= 0.5 else None

    # Parsear filas de datos
    preview_rows = []
    sin_match = []

    for row in rows[1:]:
        nombre = row[idx_nombre] if idx_nombre is not None else None
        if not nombre:
            continue
        nombre = str(nombre).strip()
        if not nombre:
            continue

        # Username: primero del Excel, luego fuzzy match
        username_excel = None
        if idx_username is not None:
            v = row[idx_username]
            if v:
                username_excel = str(v).strip() or None

        username_matched = username_excel or buscar_username(nombre)
        confianza = "manual" if username_excel else ("auto" if username_matched else "sin_match")

        horarios_dia = []
        for i, (dia_label, key_e, key_s) in enumerate(DIAS_COLS):
            entrada = _parse_hora(row[col_map[key_e]] if col_map.get(key_e) is not None else None)
            salida  = _parse_hora(row[col_map[key_s]] if col_map.get(key_s) is not None else None)
            horarios_dia.append({
                "fecha":        fechas_semana[i],
                "dia":          dia_label,
                "hora_entrada": entrada,
                "hora_salida":  salida,
            })

        entry = {
            "nombre_excel":     nombre,
            "username":         username_matched,
            "confianza":        confianza,
            "horarios":         horarios_dia,
        }
        preview_rows.append(entry)
        if not username_matched:
            sin_match.append(nombre)

    # Lista de técnicos disponibles para el selector de corrección manual
    try:
        tecnicos_disponibles = [u["username"] for u in (execute_read(
            "SELECT username FROM users WHERE role IN ('tecnico','lider') ORDER BY username"
        ) or [])]
    except Exception:
        tecnicos_disponibles = [u["username"] for u in users_db]

    return {
        "semana":               semana,
        "fechas_semana":        fechas_semana,
        "preview":              preview_rows,
        "sin_match":            sin_match,
        "tecnicos_disponibles": tecnicos_disponibles,
        "total":                len(preview_rows),
    }


# ── POST /api/horarios/confirmar-importacion ────────────────────────────────────
# Recibe el preview (ya con usernames corregidos) y los guarda en la BD.

class HorarioImportDia(BaseModel):
    fecha:        str
    hora_entrada: Optional[str] = None
    hora_salida:  Optional[str] = None

class HorarioImportRow(BaseModel):
    username: str
    horarios: List[HorarioImportDia]

class ConfirmarImportacion(BaseModel):
    semana:   str
    registros: List[HorarioImportRow]

@router.post("/confirmar-importacion")
async def confirmar_importacion(payload: ConfirmarImportacion, current_user=Depends(verify_token)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores")

    usernames_payload = list({row.username for row in payload.registros if row.username})
    fechas_payload = list({dia.fecha for row in payload.registros for dia in row.horarios})

    existentes = {}
    if usernames_payload and fechas_payload:
        ph_u = ",".join(["%s"] * len(usernames_payload))
        ph_f = ",".join(["%s"] * len(fechas_payload))
        rows = execute_read(
            f"SELECT username, fecha, hora_entrada, hora_salida FROM horarios "
            f"WHERE username IN ({ph_u}) AND fecha IN ({ph_f})",
            (*usernames_payload, *fechas_payload)
        ) or []
        for r in rows:
            existentes[(r["username"], _to_str(r["fecha"]))] = r

    guardados  = 0
    eliminados = 0
    afectados  = set()

    for row in payload.registros:
        if not row.username:
            continue
        for dia in row.horarios:
            entrada = (dia.hora_entrada or "").strip() or None
            salida  = (dia.hora_salida  or "").strip() or None

            fila_actual = existentes.get((row.username, dia.fecha))
            actual_entrada = _to_str(fila_actual["hora_entrada"]) if fila_actual else None
            actual_salida  = _to_str(fila_actual["hora_salida"])  if fila_actual else None

            if not entrada and not salida:
                if fila_actual:
                    execute_write(
                        "DELETE FROM horarios WHERE username=%s AND fecha=%s",
                        (row.username, dia.fecha)
                    )
                    eliminados += 1
                    afectados.add(row.username)
                continue

            if fila_actual and entrada == actual_entrada and salida == actual_salida:
                continue  # sin cambios: no escribir ni notificar

            if fila_actual:
                execute_write(
                    "UPDATE horarios SET hora_entrada=%s, hora_salida=%s, semana=%s "
                    "WHERE username=%s AND fecha=%s",
                    (entrada, salida, payload.semana, row.username, dia.fecha)
                )
            else:
                execute_write(
                    "INSERT INTO horarios (username, fecha, semana, hora_entrada, hora_salida) "
                    "VALUES (%s,%s,%s,%s,%s)",
                    (row.username, dia.fecha, payload.semana, entrada, salida)
                )
            guardados += 1
            afectados.add(row.username)

    if afectados:
        from routers.ws import notify
        await notify("horario_actualizado", {
            "usernames": sorted(afectados),
            "semana": payload.semana,
            "cantidad": len(afectados),
        })
        registrar_alertas_horario(sorted(afectados), payload.semana)

    return {
        "ok": True,
        "guardados": guardados,
        "eliminados": eliminados,
        "notificados": len(afectados),
    }


# ── Comentarios semanales de asistencia ─────────────────────────────────────

class ComentarioItem(BaseModel):
    username: str
    comentario: str = ""


class ComentariosBatch(BaseModel):
    semana: str
    comentarios: List[ComentarioItem]


@router.get("/comentarios")
def get_comentarios(
    semana: str = Query(...),
    current_user=Depends(verify_token)
):
    if current_user["role"] not in ("admin", "visor"):
        raise HTTPException(status_code=403, detail="Solo administradores o visores")

    rows = execute_read(
        "SELECT username, comentario FROM comentarios_asistencia WHERE semana=%s",
        (semana,)
    ) or []
    return {r["username"]: r["comentario"] for r in rows}


@router.post("/comentarios")
def save_comentarios(payload: ComentariosBatch, current_user=Depends(verify_token)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores")

    guardados = 0
    for item in payload.comentarios:
        texto = (item.comentario or "").strip()
        existente = execute_read(
            "SELECT id FROM comentarios_asistencia WHERE username=%s AND semana=%s",
            (item.username, payload.semana)
        )
        if not texto:
            if existente:
                execute_write(
                    "DELETE FROM comentarios_asistencia WHERE username=%s AND semana=%s",
                    (item.username, payload.semana)
                )
                guardados += 1
            continue
        if existente:
            execute_write(
                "UPDATE comentarios_asistencia SET comentario=%s WHERE username=%s AND semana=%s",
                (texto, item.username, payload.semana)
            )
        else:
            execute_write(
                "INSERT INTO comentarios_asistencia (username, semana, comentario) VALUES (%s,%s,%s)",
                (item.username, payload.semana, texto)
            )
        guardados += 1

    return {"ok": True, "guardados": guardados}


# ── GET /api/horarios/hoy ──────────────────────────────────────────────────────

@router.get("/hoy")
def get_horario_hoy(
    username: str = Query(...),
    fecha: str    = Query(...),
    current_user=Depends(verify_token)
):
    rows = execute_read(
        "SELECT hora_entrada, hora_salida FROM horarios "
        "WHERE username=%s AND fecha=%s LIMIT 1",
        (username, fecha)
    )
    if not rows:
        return {"horario": None}
    h = rows[0]
    return {
        "horario": {
            "hora_entrada": _to_str(h["hora_entrada"]),
            "hora_salida":  _to_str(h["hora_salida"]),
        }
    }


# ── GET /api/horarios/resumen ──────────────────────────────────────────────────
# FIX PRINCIPAL: query unificada que garantiza que técnicos con horario
# pero sin check-in aparezcan como "ausente", y que el cruce de datos
# use DATE() explícito para evitar desfases de timezone.

@router.get("/resumen")
def get_resumen(
    semana: str = Query(...),
    current_user=Depends(verify_token)
):
    if current_user["role"] not in ("admin", "visor"):
        raise HTTPException(status_code=403, detail="Solo administradores o visores")

    try:
        lunes = date.fromisoformat(semana)
    except ValueError:
        return []

    fechas = [(lunes + timedelta(days=i)).isoformat() for i in range(6)]
    placeholders = ",".join(["%s"] * len(fechas))

    # ── Horarios programados ──────────────────────────────────────────────────
    horarios_raw = execute_read(
        "SELECT username, fecha, hora_entrada, hora_salida "
        "FROM horarios WHERE semana=%s",
        (semana,)
    ) or []

    horarios_map: dict = {}
    for h in horarios_raw:
        f = _to_str(h["fecha"])
        horarios_map[(h["username"], f)] = {
            "hora_entrada": _to_str(h["hora_entrada"]),
            "hora_salida":  _to_str(h["hora_salida"]),
        }

    # ── Registros reales de check-in ──────────────────────────────────────────
    # FIX: usamos DATE(fecha) para evitar que diferencias de timezone rompan
    # el filtro, y traemos todos los campos necesarios en una sola query.
    registros_raw = execute_read(
        f"""
        SELECT username,
               DATE_FORMAT(fecha, '%%Y-%%m-%%d') AS fecha,
               tipo,
               hora_checkin,
               distancia_metros,
               aprobado,
               retardo_min
        FROM registros_asistencia
        WHERE DATE(fecha) IN ({placeholders})
        ORDER BY hora_checkin
        """,
        tuple(fechas)
    ) or []

    # reg_map[(username, fecha, tipo)] = registro
    # Si hay duplicados de tipo (dos entradas el mismo día), queda la primera.
    reg_map: dict = {}
    for r in registros_raw:
        f = str(r["fecha"])
        key = (r["username"], f, r["tipo"])
        if key not in reg_map:
            reg_map[key] = r

    # ── Técnicos: unión de los que tienen horario + los que checan aunque sea ──
    # FIX duplicados fantasma: al borrar un usuario solo se hace DELETE FROM users
    # (no hay cascada), así que sus horarios/registros viejos quedan huérfanos en
    # la BD. Antes esta lista se armaba directo de esas tablas, así que un técnico
    # ya eliminado seguía apareciendo para siempre. Ahora se cruza contra los
    # usernames que SIGUEN existiendo en `users`, sin borrar el historial.
    usuarios_vigentes = {
        u["username"] for u in (execute_read("SELECT username FROM users") or [])
    }
    tecnicos = (
        {h["username"] for h in horarios_raw}
        | {r["username"] for r in registros_raw}
    ) & usuarios_vigentes

    resultado = []
    for username in sorted(tecnicos):
        for fecha in fechas:
            horario = horarios_map.get((username, fecha))
            entrada = reg_map.get((username, fecha, "entrada"))
            salida  = reg_map.get((username, fecha, "salida"))

            # Si no hay horario programado ni registro, omitir la fila
            if not horario and not entrada and not salida:
                continue

            # Horas programadas en minutos
            he_prog = _hhmm_to_min(horario["hora_entrada"] if horario else None)
            hs_prog = _hhmm_to_min(horario["hora_salida"]  if horario else None)

            # Horas reales (hora_checkin puede ser timedelta de pymysql TIME)
            he_str = _to_str(entrada["hora_checkin"])[:5] if entrada and entrada.get("hora_checkin") else None
            hs_str = _to_str(salida["hora_checkin"])[:5]  if salida  and salida.get("hora_checkin")  else None

            he_real = _hhmm_to_min(he_str)
            hs_real = _hhmm_to_min(hs_str)

            # Retardo: preferir el valor ya calculado al momento del check-in,
            # recalcular solo si falta (migración de registros viejos).
            if entrada and entrada.get("retardo_min") is not None:
                retardo_min = int(entrada["retardo_min"])
            elif he_prog is not None and he_real is not None:
                TOLERANCIA = 15
                retardo_min = max(0, he_real - he_prog - TOLERANCIA)
            else:
                retardo_min = 0

            salida_anticipada_min = (
                max(0, hs_prog - hs_real)
                if hs_prog is not None and hs_real is not None
                else 0
            )

            horas_trabajadas = (
                round(max(0, hs_real - he_real) / 60, 2)
                if he_real is not None and hs_real is not None
                else None
            )

            # ── Estado ───────────────────────────────────────────────────────
            tiene_horario = horario and (
                horario.get("hora_entrada") or horario.get("hora_salida")
            )

            if not tiene_horario:
                estado = "libre"
            elif not entrada and not salida:
                estado = "ausente"          # ← inasistencia detectada aquí
            elif entrada and not salida:
                estado = "sin_salida"
            elif not entrada and salida:
                estado = "sin_entrada"
            else:
                estado = "completo"

            resultado.append({
                "username":             username,
                "fecha":                fecha,
                # Programado
                "hora_entrada":         horario["hora_entrada"] if horario else None,
                "hora_salida":          horario["hora_salida"]  if horario else None,
                # Real (check-in)
                "hora_entrada_real":    he_str,
                "hora_salida_real":     hs_str,
                "hora_checkin":         he_str,   # alias para compatibilidad con frontend
                # Métricas
                "retardo_min":          retardo_min,
                "salida_anticipada_min": salida_anticipada_min,
                "horas_trabajadas":     horas_trabajadas,
                # Estado
                "estado":               estado,
                # Extra para el admin
                "distancia_metros":     entrada.get("distancia_metros") if entrada else None,
                "aprobado":             bool(entrada.get("aprobado")) if entrada else None,
            })

    return resultado


# ── GET /api/horarios/resumen/excel ─────────────────────────────────────────────
# Exporta el Resumen Semanal de Asistencia como archivo .xlsx (en vez de imagen).
# Reutiliza get_resumen() y get_comentarios() para garantizar exactamente los
# mismos datos que se muestran en la tabla del navegador.

DIAS_SEMANA_XLSX = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]

ESTADO_LABEL_XLSX = {
    "completo":    "Completo",
    "retardo":     "Retardo",
    "ausente":     "Ausente",
    "libre":       "Libre",
    "sin_salida":  "Sin salida",
    "sin_entrada": "Sin entrada",
}


@router.get("/resumen/excel")
def exportar_resumen_semanal_excel(
    semana: str = Query(...),
    current_user=Depends(verify_token)
):
    if current_user["role"] not in ("admin", "visor"):
        raise HTTPException(status_code=403, detail="Solo administradores o visores")

    try:
        lunes = date.fromisoformat(semana)
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de semana inválido, debe ser YYYY-MM-DD")

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no instalado. Agrega 'openpyxl' a requirements.txt")

    resumen = get_resumen(semana=semana, current_user=current_user)
    comentarios = get_comentarios(semana=semana, current_user=current_user)

    fechas = [(lunes + timedelta(days=i)).isoformat() for i in range(6)]

    # ── username → nombre_completo (nombre real del técnico) ──────────────────
    # Igual patrón usado en reporte_router.py / dashboard_router.py: si el técnico
    # no tiene nombre_completo capturado, cae de vuelta a su username.
    filas_nombres = execute_read("SELECT username, nombre_completo FROM users") or []
    nombres_map = {
        r["username"]: (r["nombre_completo"] or r["username"]) for r in filas_nombres
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen Semanal"

    header_fill = PatternFill(start_color="002B5B", end_color="002B5B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ESTADO_FILL = {
        "completo":    PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
        "retardo":     PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "ausente":     PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "libre":       PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),
        "sin_salida":  PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "sin_entrada": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    }

    # ── Encabezados ──────────────────────────────────────────────────────────
    headers = ["Técnico"]
    for i, f in enumerate(fechas):
        headers.append(f"{DIAS_SEMANA_XLSX[i]}\n{f[5:]}")
    headers += ["Hrs. trabajadas", "Horas extra", "Retardos", "Comentarios"]

    for col, titulo in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=titulo)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    # ── Datos ────────────────────────────────────────────────────────────────
    tecnicos = sorted(
        {r["username"] for r in resumen},
        key=lambda u: nombres_map.get(u, u).lower()
    )
    row_idx = 2
    for tec in tecnicos:
        filas = [r for r in resumen if r["username"] == tec]
        total_hrs = 0.0
        num_retardos = 0
        min_retardos = 0
        for r in filas:
            if r.get("retardo_min"):
                if r["retardo_min"] > 0:
                    num_retardos += 1
                    min_retardos += r["retardo_min"]

        col = 1
        c = ws.cell(row=row_idx, column=col, value=nombres_map.get(tec, tec))
        c.border = border
        c.alignment = left
        c.font = Font(bold=True)

        for f in fechas:
            col += 1
            r = next((x for x in filas if x["fecha"] == f), None)
            if not r:
                c = ws.cell(row=row_idx, column=col, value="Libre")
                c.fill = ESTADO_FILL["libre"]
            else:
                if r.get("horas_trabajadas"):
                    total_hrs += float(r["horas_trabajadas"])
                estado = r["estado"]
                label = ESTADO_LABEL_XLSX.get(estado, estado)
                detalle = ""
                if r.get("hora_entrada_real") or r.get("hora_salida_real"):
                    detalle = f"\n{r.get('hora_entrada_real') or '—'} → {r.get('hora_salida_real') or '—'}"
                ret_txt = f"\n+{r['retardo_min']}min retardo" if r.get("retardo_min") and r["retardo_min"] > 0 else ""
                c = ws.cell(row=row_idx, column=col, value=f"{label}{detalle}{ret_txt}")
                c.fill = ESTADO_FILL.get(estado, ESTADO_FILL["libre"])
            c.border = border
            c.alignment = center

        horas_extra = max(0.0, total_hrs - 48)
        col += 1
        c = ws.cell(row=row_idx, column=col, value=round(total_hrs, 1))
        c.border = border
        c.alignment = center
        c.font = Font(bold=True)

        col += 1
        c = ws.cell(row=row_idx, column=col, value=round(horas_extra, 1) if horas_extra > 0 else "—")
        c.border = border
        c.alignment = center
        if horas_extra > 0:
            c.font = Font(color="D97706", bold=True)

        col += 1
        if num_retardos > 0:
            c = ws.cell(row=row_idx, column=col, value=f"{num_retardos} ({min_retardos} min)")
            c.fill = ESTADO_FILL["retardo"]
        else:
            c = ws.cell(row=row_idx, column=col, value="—")
        c.border = border
        c.alignment = center

        col += 1
        c = ws.cell(row=row_idx, column=col, value=comentarios.get(tec, "") or "")
        c.border = border
        c.alignment = left

        row_idx += 1

    # ── Anchos de columna ────────────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(1)].width = 18
    for i in range(2, 2 + len(fechas)):
        ws.column_dimensions[get_column_letter(i)].width = 20
    ws.column_dimensions[get_column_letter(2 + len(fechas))].width = 16
    ws.column_dimensions[get_column_letter(3 + len(fechas))].width = 14
    ws.column_dimensions[get_column_letter(4 + len(fechas))].width = 16
    ws.column_dimensions[get_column_letter(5 + len(fechas))].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"resumen_semanal_asistencia_{semana}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
